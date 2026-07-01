from __future__ import annotations

import base64
import copy
import json
import logging
import math
import os
import re
import shutil
import subprocess
import sys
import threading
import traceback
import uuid
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib import error, request

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception:  # pragma: no cover - import guard for friendly CLI errors
    tk = None  # type: ignore[assignment]
    filedialog = None  # type: ignore[assignment]
    messagebox = None  # type: ignore[assignment]
    ttk = None  # type: ignore[assignment]

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.views import Selection
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]
    MergedCell = None  # type: ignore[assignment]
    XLImage = None  # type: ignore[assignment]
    Alignment = None  # type: ignore[assignment]
    Font = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]
    Selection = None  # type: ignore[assignment]

try:
    from PIL import Image, ImageOps, ImageTk
except Exception:  # pragma: no cover
    Image = None  # type: ignore[assignment]
    ImageOps = None  # type: ignore[assignment]
    ImageTk = None  # type: ignore[assignment]


APP_DIR = Path(__file__).resolve().parent
CONFIG_DIR = APP_DIR / "config"
TEMPLATE_DIR = APP_DIR / "templates"
OUTPUT_DIR = APP_DIR / "outputs"
WORK_DIR = APP_DIR / "work"
LOG_DIR = APP_DIR / "logs"
UNPROCESSED_DIR = APP_DIR / "Unprocessed"
PROCESSED_DIR = APP_DIR / "Processed"
USER_SETTINGS_FILE = CONFIG_DIR / "user_settings.json"
USER_API_KEY_FILE = CONFIG_DIR / "api_key.txt"
TEMPLATES_CONFIG_FILE = CONFIG_DIR / "templates.json"
SESSION_FILE = CONFIG_DIR / "session_state.json"
SUPPORTED_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
SUPPORTED_UPLOAD_SUFFIXES = SUPPORTED_IMAGE_SUFFIXES | {".pdf"}
OPENAI_CHAT_COMPLETIONS_URL = "https://api.openai.com/v1/chat/completions"
DEFAULT_OPENAI_MODEL = "gpt-4o-mini"
DEFAULT_USD_TO_RMB_RATE = "6.8175"
DEFAULT_USD_TO_KRW_RATE = "1548.86"
DEFAULT_KRW_TO_RMB_RATE = "0.004433"
KRW_NUMBER_FORMAT = "\u20a9#,##0"
LOGGER = logging.getLogger("reimbursement_helper")


USA_TEMPLATE_NAME = "usa_expense_report_template.xlsx"
KOREA_COVER_TEMPLATE_NAME = "korea_cover_receipts_template.xlsx"
KOREA_DETAILS_TEMPLATE_NAME = "korea_details_template.xlsx"


def setup_logging() -> None:
    if LOGGER.handlers:
        return
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    handler = RotatingFileHandler(
        LOG_DIR / "app.log",
        maxBytes=750_000,
        backupCount=5,
        encoding="utf-8",
    )
    handler.setFormatter(
        logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    )
    LOGGER.setLevel(logging.INFO)
    LOGGER.addHandler(handler)
    LOGGER.propagate = False
    LOGGER.info("Reimbursement Helper logging started")


def log_exception(message: str) -> Path:
    setup_logging()
    LOGGER.exception(message)
    return LOG_DIR / "app.log"


def ensure_runtime_folders() -> None:
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    UNPROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)


def supported_image_files(folder: Path) -> List[Path]:
    if not folder.exists():
        return []
    return sorted(
        [
            path
            for path in folder.iterdir()
            if path.is_file() and path.suffix.lower() in SUPPORTED_IMAGE_SUFFIXES
        ],
        key=lambda path: path.name.lower(),
    )


def render_pdf_pages(pdf_path: Path) -> List[Path]:
    try:
        import fitz  # type: ignore[import-not-found]
    except Exception as exc:
        raise RuntimeError(
            "PDF support requires PyMuPDF. Run `python -m pip install -r requirements.txt`, then try again."
        ) from exc

    target_dir = WORK_DIR / "pdf_pages" / f"{pdf_path.stem}_{uuid.uuid4().hex[:8]}"
    target_dir.mkdir(parents=True, exist_ok=True)
    rendered: List[Path] = []
    try:
        document = fitz.open(str(pdf_path))
    except Exception as exc:
        raise RuntimeError(f"Could not open PDF: {pdf_path}") from exc
    try:
        for page_index in range(document.page_count):
            page = document.load_page(page_index)
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            output = target_dir / f"{pdf_path.stem}_page_{page_index + 1}.png"
            pixmap.save(str(output))
            rendered.append(output)
    finally:
        document.close()
    return rendered


def render_pdf_to_merged_image(pdf_path: Path) -> Path:
    page_paths = render_pdf_pages(pdf_path)
    if not page_paths:
        raise RuntimeError(f"PDF has no pages: {pdf_path}")
    if Image is None:
        raise RuntimeError(
            "PDF merge requires Pillow. Run `python -m pip install -r requirements.txt`, then try again."
        )
    target_dir = WORK_DIR / "pdf_pages" / f"{pdf_path.stem}_{uuid.uuid4().hex[:8]}"
    target_dir.mkdir(parents=True, exist_ok=True)
    images = []
    try:
        for page_path in page_paths:
            with Image.open(page_path) as image:
                images.append(image.convert("RGB").copy())
        max_width = max(image.width for image in images)
        gutter = 24
        total_height = sum(image.height for image in images) + gutter * (len(images) - 1)
        merged = Image.new("RGB", (max_width, total_height), "#FFFFFF")
        y = 0
        for image in images:
            x = (max_width - image.width) // 2
            merged.paste(image, (x, y))
            y += image.height + gutter
        output = target_dir / f"{pdf_path.stem}_merged.png"
        merged.save(output, "PNG")
        return output
    finally:
        for image in images:
            try:
                image.close()
            except Exception:
                pass


def selected_file_key(path: Path, source_path: str = "", source_page: str = "") -> Tuple[str, str]:
    key_path = source_path or str(path)
    try:
        key_path = str(Path(key_path).resolve()).lower()
    except Exception:
        key_path = str(key_path).lower()
    return key_path, str(source_page or "")


def unique_path(folder: Path, filename: str) -> Path:
    candidate = folder / filename
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    for index in range(2, 10_000):
        candidate = folder / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
    return folder / f"{stem}_{uuid.uuid4().hex}{suffix}"


def same_folder(path: Path, folder: Path) -> bool:
    try:
        return path.resolve().parent == folder.resolve()
    except Exception:
        return False


def open_path(path: Path) -> None:
    if sys.platform == "win32":
        os.startfile(str(path))  # type: ignore[attr-defined]
    else:
        raise RuntimeError(f"Opening files is only implemented for Windows. Path: {path}")


def reveal_in_file_explorer(path: Path) -> None:
    if sys.platform == "win32":
        resolved = path.resolve()
        if resolved.exists():
            subprocess.Popen(["explorer.exe", f'/select,"{resolved}"'])
        else:
            open_path(resolved.parent)
    else:
        open_path(path.parent)


class ToolTip:
    def __init__(self, widget: Any, text: str) -> None:
        self.widget = widget
        self.text = text
        self.tip_window: Optional[Any] = None
        self.after_id: Optional[str] = None
        widget.bind("<Enter>", self.schedule)
        widget.bind("<Leave>", self.hide)
        widget.bind("<ButtonPress>", self.hide)

    def schedule(self, _event: Any = None) -> None:
        self.cancel()
        try:
            self.after_id = self.widget.after(450, self.show)
        except Exception:
            self.after_id = None

    def cancel(self) -> None:
        if self.after_id:
            try:
                self.widget.after_cancel(self.after_id)
            except Exception:
                pass
            self.after_id = None

    def show(self) -> None:
        if self.tip_window is not None or tk is None:
            return
        try:
            x = self.widget.winfo_rootx() + 18
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 8
        except Exception:
            return
        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)
        self.tip_window.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            self.tip_window,
            text=self.text,
            background="#111827",
            foreground="#F8FAFC",
            borderwidth=0,
            padx=8,
            pady=4,
            font=("Segoe UI", 9),
        )
        label.pack()

    def hide(self, _event: Any = None) -> None:
        self.cancel()
        if self.tip_window is not None:
            try:
                self.tip_window.destroy()
            except Exception:
                pass
            self.tip_window = None


USA_CATEGORY_ROWS: Dict[str, List[int]] = {
    "transportation": list(range(7, 47)),
    "lodging": list(range(49, 52)),
    "meals": [54],
    "advertising": [57],
    "office": list(range(60, 65)),
    "entertainment": list(range(67, 69)),
    "other": list(range(60, 65)),
}

KOREA_CATEGORY_COLUMNS: Dict[str, str] = {
    "transportation": "F",
    "physical_exam": "G",
    "lodging": "H",
    "nucleic_test": "I",
    "materials": "J",
    "meals": "K",
    "courier": "L",
    "consumables": "M",
    "welfare": "N",
    "other": "O",
}

USA_REPORT_CATEGORY_ORDER = [
    "transportation",
    "lodging",
    "meals",
    "advertising",
    "office",
    "entertainment",
    "other",
]

KOREA_REPORT_CATEGORY_ORDER = [
    "transportation",
    "physical_exam",
    "lodging",
    "nucleic_test",
    "materials",
    "meals",
    "courier",
    "consumables",
    "welfare",
    "other",
]

KOREA_COVER_ROWS: Dict[str, Tuple[int, str]] = {
    "transportation": (4, "交通费"),
    "consumables": (5, "消耗品"),
    "lodging": (6, "住宿费"),
    "meals": (7, "业务招待费/餐费"),
    "other": (8, "其他"),
}

USA_CATEGORY_LABELS = {
    "transportation": "Transportation",
    "lodging": "Lodging",
    "meals": "Meals",
    "advertising": "Advertising",
    "office": "Office",
    "entertainment": "Entertainment",
    "other": "Other",
}

KOREA_CATEGORY_LABELS = {
    "transportation": "Transportation / 交通费",
    "physical_exam": "Physical exam / 入职体检费",
    "lodging": "Accommodation / 住宿费",
    "nucleic_test": "Nucleic acid test / 核酸检测费",
    "materials": "Material / 物料费",
    "meals": "Meals / 业务招待费-餐费",
    "courier": "Courier / 快递费",
    "consumables": "Consumables / 消耗品",
    "welfare": "Welfare / 福利费",
    "other": "Other / 其他",
}


KOREA_COVER_ROWS = {
    "transportation": (4, "\u4ea4\u901a\u8d39"),
    "consumables": (5, "\u6d88\u8017\u54c1"),
    "lodging": (6, "\u4f4f\u5bbf\u8d39"),
    "meals": (7, "\u4e1a\u52a1\u62db\u5f85\u8d39/\u9910\u8d39"),
    "other": (8, "\u5176\u4ed6"),
}

KOREA_CATEGORY_LABELS = {
    "transportation": "Transportation / \u4ea4\u901a\u8d39",
    "physical_exam": "Physical exam / \u5165\u804c\u4f53\u68c0\u8d39",
    "lodging": "Accommodation / \u4f4f\u5bbf\u8d39",
    "nucleic_test": "Nucleic acid test / \u6838\u9178\u68c0\u6d4b\u8d39",
    "materials": "Material / \u7269\u6599\u8d39",
    "meals": "Meals / \u4e1a\u52a1\u62db\u5f85\u8d39-\u9910\u8d39",
    "courier": "Courier / \u5feb\u9012\u8d39",
    "consumables": "Consumables / \u6d88\u8017\u54c1",
    "welfare": "Welfare / \u798f\u5229\u8d39",
    "other": "Other / \u5176\u4ed6",
}

CATEGORY_LABELS_BY_LANGUAGE: Dict[str, Dict[str, str]] = {
    "en": {
        "transportation": "Transportation",
        "lodging": "Lodging",
        "meals": "Meals",
        "advertising": "Advertising",
        "office": "Office",
        "entertainment": "Entertainment",
        "materials": "Materials",
        "consumables": "Consumables",
        "physical_exam": "Physical exam",
        "nucleic_test": "Nucleic acid test",
        "courier": "Courier",
        "welfare": "Welfare",
        "other": "Other",
    },
    "zh": {
        "transportation": "\u4ea4\u901a\u8d39",
        "lodging": "\u4f4f\u5bbf\u8d39",
        "meals": "\u4e1a\u52a1\u62db\u5f85\u8d39-\u9910\u8d39",
        "advertising": "\u5e7f\u544a\u8d39",
        "office": "\u529e\u516c\u8d39",
        "entertainment": "\u4e1a\u52a1\u62db\u5f85\u8d39",
        "materials": "\u7269\u6599\u8d39",
        "consumables": "\u6d88\u8017\u54c1",
        "physical_exam": "\u5165\u804c\u4f53\u68c0\u8d39",
        "nucleic_test": "\u6838\u9178\u68c0\u6d4b\u8d39",
        "courier": "\u5feb\u9012\u8d39",
        "welfare": "\u798f\u5229\u8d39",
        "other": "\u5176\u4ed6",
    },
    "ko": {
        "transportation": "\uad50\ud1b5\ube44",
        "lodging": "\uc219\ubc15\ube44",
        "meals": "\uc2dd\ub300",
        "advertising": "\uad11\uace0\ube44",
        "office": "\uc0ac\ubb34\ube44",
        "entertainment": "\uc811\ub300\ube44",
        "materials": "\uc790\uc7ac\ube44",
        "consumables": "\uc18c\ubaa8\ud488",
        "physical_exam": "\uac74\uac15\uac80\uc9c4\ube44",
        "nucleic_test": "\ud575\uc0b0 \uac80\uc0ac\ube44",
        "courier": "\ud0dd\ubc30\ube44",
        "welfare": "\ubcf5\ub9ac\ud6c4\uc0dd\ube44",
        "other": "\uae30\ud0c0",
    },
}

LANGUAGE_LABELS = {
    "en": "English",
    "zh": "\u4e2d\u6587",
    "ko": "\ud55c\uad6d\uc5b4",
}

UI_TEXT: Dict[str, Dict[str, str]] = {
    "en": {
        "app_title": "Reimbursement Helper",
        "header_title": "REIMBURSEMENT HELPER",
        "form": "Form",
        "language": "Language",
        "select_files": "Select Files",
        "select_payment_proof": "Select Payment Proof",
        "select_exchange_rate": "Select Exchange Rate Images",
        "generate_details": "Generate Details",
        "generate_all": "Generate All",
        "generate_excel": "Generate Excel",
        "inserted_receipts": "Inserted receipts and details",
        "file": "File",
        "status": "Status",
        "date": "Date",
        "amount": "Amount",
        "remove": "Remove",
        "clear": "Clear",
        "details": "Details",
        "receipt_preview": "Receipt preview",
        "rotate_left": "Rotate left",
        "rotate_right": "Rotate right",
        "revert": "Revert",
        "delete_screenshot": "Delete selected screenshot",
        "open_output_folder": "Open output folder",
        "ready": "Ready",
        "place_vendor": "Place / Vendor",
        "usd_amount": "USD amount",
        "original_amount": "Original amount",
        "currency": "Currency",
        "krw_amount": "KRW amount",
        "rmb_amount": "RMB amount",
        "purpose": "Purpose",
        "project_number": "Project number",
        "category": "Category",
        "payment_method": "Payment method",
        "receipt_label": "Receipt label",
        "select_receipt_prompt": "Select receipt image or PDF files to begin.",
        "proof_prompt": "Select Payment Proof, then Generate All.",
        "no_receipt_screenshot": "No receipt screenshot.",
        "exchange_only_korea": "Exchange-rate images are only used for the Korea form.",
        "selected_exchange_files": "Selected {count} exchange-rate image file(s).",
        "selected_exchange_no_key": "Selected exchange-rate image(s). Add an API key to auto-read the rate.",
        "reading_exchange": "Reading exchange-rate image...",
        "reading_exchange_short": "Reading exchange rate",
        "selected_exchange_failed": "Selected exchange-rate image(s), but could not auto-read rate: {message}",
        "updated_exchange_rate": "Updated rate from exchange-rate image: {updates}",
    },
    "zh": {
        "app_title": "\u62a5\u9500\u52a9\u624b",
        "header_title": "\u62a5\u9500\u52a9\u624b",
        "form": "\u8868\u683c",
        "language": "\u8bed\u8a00",
        "select_files": "\u9009\u62e9\u6587\u4ef6",
        "select_payment_proof": "\u9009\u62e9\u4ed8\u6b3e\u51ed\u8bc1",
        "select_exchange_rate": "\u9009\u62e9\u6c47\u7387\u56fe\u7247",
        "generate_details": "\u751f\u6210\u660e\u7ec6",
        "generate_all": "\u5168\u90e8\u751f\u6210",
        "generate_excel": "\u751f\u6210 Excel",
        "inserted_receipts": "\u5df2\u63d2\u5165\u53d1\u7968\u548c\u660e\u7ec6",
        "file": "\u6587\u4ef6",
        "status": "\u72b6\u6001",
        "date": "\u65e5\u671f",
        "amount": "\u91d1\u989d",
        "remove": "\u5220\u9664",
        "clear": "\u6e05\u7a7a",
        "details": "\u660e\u7ec6",
        "receipt_preview": "\u53d1\u7968\u9884\u89c8",
        "rotate_left": "\u5411\u5de6\u65cb\u8f6c",
        "rotate_right": "\u5411\u53f3\u65cb\u8f6c",
        "revert": "\u8fd8\u539f",
        "delete_screenshot": "\u5220\u9664\u9009\u4e2d\u622a\u56fe",
        "open_output_folder": "\u6253\u5f00\u8f93\u51fa\u6587\u4ef6\u5939",
        "ready": "\u5c31\u7eea",
        "place_vendor": "\u5730\u70b9 / \u5546\u5bb6",
        "usd_amount": "\u7f8e\u5143\u91d1\u989d",
        "original_amount": "\u539f\u5e01\u91d1\u989d",
        "currency": "\u5e01\u79cd",
        "krw_amount": "\u97e9\u5143\u91d1\u989d",
        "rmb_amount": "\u4eba\u6c11\u5e01\u91d1\u989d",
        "purpose": "\u7528\u9014",
        "project_number": "\u9879\u76ee\u53f7",
        "category": "\u7c7b\u522b",
        "payment_method": "\u4ed8\u6b3e\u65b9\u5f0f",
        "receipt_label": "\u53d1\u7968\u6807\u7b7e",
        "select_receipt_prompt": "\u8bf7\u9009\u62e9\u53d1\u7968\u56fe\u7247\u6216 PDF \u6587\u4ef6\u3002",
        "proof_prompt": "\u8bf7\u5148\u9009\u62e9\u4ed8\u6b3e\u51ed\u8bc1\uff0c\u7136\u540e\u5168\u90e8\u751f\u6210\u3002",
        "no_receipt_screenshot": "\u6ca1\u6709\u53d1\u7968\u622a\u56fe\u3002",
        "exchange_only_korea": "\u6c47\u7387\u56fe\u7247\u4ec5\u7528\u4e8e\u97e9\u56fd\u8868\u683c\u3002",
        "selected_exchange_files": "\u5df2\u9009\u62e9 {count} \u4e2a\u6c47\u7387\u56fe\u7247\u6587\u4ef6\u3002",
        "selected_exchange_no_key": "\u5df2\u9009\u62e9\u6c47\u7387\u56fe\u7247\u3002\u6dfb\u52a0 API \u5bc6\u94a5\u540e\u53ef\u81ea\u52a8\u8bfb\u53d6\u6c47\u7387\u3002",
        "reading_exchange": "\u6b63\u5728\u8bfb\u53d6\u6c47\u7387\u56fe\u7247...",
        "reading_exchange_short": "\u6b63\u5728\u8bfb\u53d6\u6c47\u7387",
        "selected_exchange_failed": "\u5df2\u9009\u62e9\u6c47\u7387\u56fe\u7247\uff0c\u4f46\u65e0\u6cd5\u81ea\u52a8\u8bfb\u53d6\u6c47\u7387\uff1a{message}",
        "updated_exchange_rate": "\u5df2\u6839\u636e\u6c47\u7387\u56fe\u7247\u66f4\u65b0\uff1a{updates}",
    },
    "ko": {
        "app_title": "\ud658\uae09 \ub3c4\uc6b0\ubbf8",
        "header_title": "\ud658\uae09 \ub3c4\uc6b0\ubbf8",
        "form": "\uc591\uc2dd",
        "language": "\uc5b8\uc5b4",
        "select_files": "\ud30c\uc77c \uc120\ud0dd",
        "select_payment_proof": "\uacb0\uc81c \uc99d\ube59 \uc120\ud0dd",
        "select_exchange_rate": "\ud658\uc728 \uc774\ubbf8\uc9c0 \uc120\ud0dd",
        "generate_details": "\uc138\ubd80 \uc815\ubcf4 \uc0dd\uc131",
        "generate_all": "\uc804\uccb4 \uc0dd\uc131",
        "generate_excel": "Excel \uc0dd\uc131",
        "inserted_receipts": "\ucd94\uac00\ub41c \uc601\uc218\uc99d \ubc0f \uc138\ubd80 \uc815\ubcf4",
        "file": "\ud30c\uc77c",
        "status": "\uc0c1\ud0dc",
        "date": "\ub0a0\uc9dc",
        "amount": "\uae08\uc561",
        "remove": "\uc0ad\uc81c",
        "clear": "\ucd08\uae30\ud654",
        "details": "\uc138\ubd80 \uc815\ubcf4",
        "receipt_preview": "\uc601\uc218\uc99d \ubbf8\ub9ac\ubcf4\uae30",
        "rotate_left": "\uc67c\ucabd \ud68c\uc804",
        "rotate_right": "\uc624\ub978\ucabd \ud68c\uc804",
        "revert": "\ub418\ub3cc\ub9ac\uae30",
        "delete_screenshot": "\uc120\ud0dd\ud55c \uc2a4\ud06c\ub9b0\uc0f7 \uc0ad\uc81c",
        "open_output_folder": "\ucd9c\ub825 \ud3f4\ub354 \uc5f4\uae30",
        "ready": "\uc900\ube44\ub428",
        "place_vendor": "\uc7a5\uc18c / \uc5c5\uccb4",
        "usd_amount": "USD \uae08\uc561",
        "original_amount": "\uc6d0\ud654 \uae08\uc561",
        "currency": "\ud1b5\ud654",
        "krw_amount": "KRW \uae08\uc561",
        "rmb_amount": "RMB \uae08\uc561",
        "purpose": "\ubaa9\uc801",
        "project_number": "\ud504\ub85c\uc81d\ud2b8 \ubc88\ud638",
        "category": "\ubd84\ub958",
        "payment_method": "\uacb0\uc81c \uc218\ub2e8",
        "receipt_label": "\uc601\uc218\uc99d \ub77c\ubca8",
        "select_receipt_prompt": "\uc601\uc218\uc99d \uc774\ubbf8\uc9c0 \ub610\ub294 PDF \ud30c\uc77c\uc744 \uc120\ud0dd\ud558\uc138\uc694.",
        "proof_prompt": "\uacb0\uc81c \uc99d\ube59\uc744 \uc120\ud0dd\ud55c \ub4a4 \uc804\uccb4 \uc0dd\uc131\uc744 \uc2e4\ud589\ud558\uc138\uc694.",
        "no_receipt_screenshot": "\uc601\uc218\uc99d \uc2a4\ud06c\ub9b0\uc0f7\uc774 \uc5c6\uc2b5\ub2c8\ub2e4.",
        "exchange_only_korea": "\ud658\uc728 \uc774\ubbf8\uc9c0\ub294 \ud55c\uad6d \uc591\uc2dd\uc5d0\ub9cc \uc0ac\uc6a9\ub429\ub2c8\ub2e4.",
        "selected_exchange_files": "\ud658\uc728 \uc774\ubbf8\uc9c0 \ud30c\uc77c {count}\uac1c\ub97c \uc120\ud0dd\ud588\uc2b5\ub2c8\ub2e4.",
        "selected_exchange_no_key": "\ud658\uc728 \uc774\ubbf8\uc9c0\ub97c \uc120\ud0dd\ud588\uc2b5\ub2c8\ub2e4. API \ud0a4\ub97c \ucd94\uac00\ud558\uba74 \uc790\ub3d9\uc73c\ub85c \ud658\uc728\uc744 \uc77d\uc2b5\ub2c8\ub2e4.",
        "reading_exchange": "\ud658\uc728 \uc774\ubbf8\uc9c0\ub97c \uc77d\ub294 \uc911...",
        "reading_exchange_short": "\ud658\uc728 \uc77d\ub294 \uc911",
        "selected_exchange_failed": "\ud658\uc728 \uc774\ubbf8\uc9c0\ub97c \uc120\ud0dd\ud588\uc9c0\ub9cc \uc790\ub3d9\uc73c\ub85c \uc77d\uc9c0 \ubabb\ud588\uc2b5\ub2c8\ub2e4: {message}",
        "updated_exchange_rate": "\ud658\uc728 \uc774\ubbf8\uc9c0\uc5d0\uc11c \ud658\uc728\uc744 \uc5c5\ub370\uc774\ud2b8\ud588\uc2b5\ub2c8\ub2e4: {updates}",
    },
}


@dataclass
class ReceiptItem:
    item_id: str
    path: str
    filename: str
    source_path: str = ""
    source_page: str = ""
    date: str = ""
    place: str = ""
    amount: str = ""
    currency: str = "USD"
    krw_amount: str = ""
    rmb_amount: str = ""
    purpose: str = ""
    details: str = ""
    project_number: str = ""
    category: str = "transportation"
    payment_method: str = ""
    receipt_label: str = ""
    status: str = "Empty"
    crop_box: Optional[List[float]] = None
    rotation_degrees: int = 0
    receipt_images: List[Dict[str, Any]] = field(default_factory=list)


@dataclass
class BankStatementItem:
    item_id: str
    path: str
    filename: str
    source_path: str = ""
    source_page: str = ""
    date: str = ""
    amount: str = ""
    place: str = ""
    matched_receipt_id: str = ""
    status: str = "Needs AI"
    crop_box: Optional[List[float]] = None
    rotation_degrees: int = 0


def attachment_from_item(item: Any) -> Dict[str, Any]:
    return {
        "id": getattr(item, "item_id", uuid.uuid4().hex),
        "path": getattr(item, "path", ""),
        "filename": getattr(item, "filename", ""),
        "source_path": getattr(item, "source_path", ""),
        "source_page": getattr(item, "source_page", ""),
        "crop_box": copy.deepcopy(getattr(item, "crop_box", None)),
        "rotation_degrees": normalize_rotation(getattr(item, "rotation_degrees", 0)),
    }


def normalize_attachment(raw: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(raw, dict):
        return None
    path = str(raw.get("path") or "").strip()
    if not path:
        return None
    return {
        "id": str(raw.get("id") or uuid.uuid4().hex),
        "path": path,
        "filename": str(raw.get("filename") or Path(path).name),
        "source_path": str(raw.get("source_path") or ""),
        "source_page": str(raw.get("source_page") or ""),
        "crop_box": raw.get("crop_box") if isinstance(raw.get("crop_box"), list) else None,
        "rotation_degrees": normalize_rotation(raw.get("rotation_degrees", 0)),
    }


def appdata_daily_logger_key_file() -> Optional[Path]:
    appdata = os.getenv("APPDATA", "").strip()
    if not appdata:
        return None
    return Path(appdata) / "DailyLogger" / "settings" / "daily_logger_api_key.txt"


def load_user_settings() -> Dict[str, Any]:
    if not USER_SETTINGS_FILE.exists():
        return {}
    try:
        data = json.loads(USER_SETTINGS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def save_user_settings(settings: Dict[str, Any]) -> None:
    try:
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        USER_SETTINGS_FILE.write_text(
            json.dumps(settings, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        log_exception("Could not save user settings")


def normalized_form_version(value: Any) -> str:
    return "Korea" if str(value) == "Korea" else "USA"


def normalize_language(value: Any) -> str:
    return str(value) if str(value) in UI_TEXT else "en"


def ui_text(language: str, key: str, **values: Any) -> str:
    template = UI_TEXT.get(normalize_language(language), UI_TEXT["en"]).get(key, UI_TEXT["en"].get(key, key))
    for name, value in values.items():
        template = template.replace("{" + name + "}", str(value))
    return template


def load_templates_config() -> Dict[str, Any]:
    if not TEMPLATES_CONFIG_FILE.exists():
        return {}
    try:
        data = json.loads(TEMPLATES_CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def parse_env_file_for_key(path: Path) -> str:
    if not path.exists():
        return ""
    try:
        for line in path.read_text(encoding="utf-8").splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            if stripped.startswith("OPENAI_API_KEY="):
                return stripped.split("=", 1)[1].strip().strip('"').strip("'")
    except Exception:
        return ""
    return ""


def get_openai_api_key() -> str:
    key = os.getenv("OPENAI_API_KEY", "").strip()
    if key:
        return key
    key = parse_env_file_for_key(APP_DIR / ".env.local") or parse_env_file_for_key(APP_DIR / ".env")
    if key:
        return key
    if USER_API_KEY_FILE.exists():
        try:
            key = USER_API_KEY_FILE.read_text(encoding="utf-8").strip()
            if key:
                return key
        except Exception:
            pass
    daily_logger_key = appdata_daily_logger_key_file()
    if daily_logger_key and daily_logger_key.exists():
        try:
            key = daily_logger_key.read_text(encoding="utf-8").strip()
            if key:
                return key
        except Exception:
            pass
    return ""


def image_mime_for_path(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if suffix == ".webp":
        return "image/webp"
    if suffix == ".gif":
        return "image/gif"
    if suffix == ".bmp":
        return "image/bmp"
    return "image/png"


def image_data_url(path: Path) -> str:
    raw = path.read_bytes()
    return f"data:{image_mime_for_path(path)};base64,{base64.b64encode(raw).decode('ascii')}"


def extract_json_object(text: str) -> Dict[str, Any]:
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?", "", cleaned, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        parsed = json.loads(cleaned)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        pass
    match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
    if not match:
        return {}
    try:
        parsed = json.loads(match.group(0))
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def call_openai_receipt_extraction(
    image_path: Path,
    form_version: str,
    existing: ReceiptItem,
    progress: Optional[Callable[[str], None]] = None,
) -> Dict[str, Any]:
    api_key = get_openai_api_key()
    if not api_key:
        raise RuntimeError(
            "No OpenAI API key was found. Add one to .env.local, config/api_key.txt, "
            "or Daily Logger settings, then try again."
        )

    def notify(message: str) -> None:
        if progress:
            try:
                progress(message)
            except Exception:
                pass

    notify("Reading receipt image")
    prepared_image = prepare_receipt_image_file(
        image_path,
        existing.crop_box,
        existing.rotation_degrees,
    )
    data_url = image_data_url(prepared_image)
    form_hint = "USA VisionNav reimbursement form" if form_version == "USA" else "Korea VisionNav reimbursement form"
    prompt = f"""
Extract reimbursement fields from this receipt image for a {form_hint}.
Return only compact JSON with these keys:
date, place, amount, currency, krw_amount, rmb_amount, purpose, details,
project_number, category, payment_method, receipt_label, confidence_notes.

Rules:
- date must be yyyy-mm-dd when visible, otherwise empty string.
- amount should be the paid total. Preserve cents when visible.
- currency must match the receipt symbol/text: USD for "$", KRW for won/KRW,
  and RMB/CNY for yuan/RMB/CNY. Do not default to KRW just because the form is Korea.
- category must be one of:
  transportation, lodging, meals, advertising, office, entertainment, physical_exam,
  nucleic_test, materials, courier, consumables, welfare, other.
- project_number should stay empty unless printed on the receipt or obvious from context.
- purpose should be short, for example Parking, Gas, Hotel, Meal, Baggage, Supplies.
- details should be a natural one-line description.
- receipt_label should be short enough to fit above a receipt image in Excel.
- If a field is unclear, use an empty string rather than guessing.

Current editable row values, to keep if they look more specific than the image:
{json.dumps(asdict(existing), ensure_ascii=False)}
""".strip()
    payload = json.dumps(
        {
            "model": os.getenv("OPENAI_MODEL", DEFAULT_OPENAI_MODEL),
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "You extract structured receipt data for reimbursement forms. "
                        "You never invent project numbers, names, or hidden values."
                    ),
                },
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": data_url}},
                    ],
                },
            ],
            "temperature": 0,
        }
    ).encode("utf-8")
    req = request.Request(
        OPENAI_CHAT_COMPLETIONS_URL,
        data=payload,
        method="POST",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
    )
    notify("Contacting AI")
    try:
        with request.urlopen(req, timeout=90) as response:
            body = response.read().decode("utf-8")
    except error.HTTPError as exc:
        try:
            details = exc.read().decode("utf-8")
        except Exception:
            details = str(exc)
        raise RuntimeError(f"OpenAI API error ({exc.code}): {details}") from exc
    except Exception as exc:
        raise RuntimeError(f"Could not contact OpenAI: {exc}") from exc

    notify("Applying extracted fields")
    try:
        parsed = json.loads(body)
        content = parsed["choices"][0]["message"]["content"]
    except Exception as exc:
        raise RuntimeError("OpenAI returned an unexpected response format.") from exc
    data = extract_json_object(str(content))
    if not data:
        raise RuntimeError("AI response did not contain usable JSON.")
    return data


def exchange_rates_from_data(data: Dict[str, Any], usd_to_rmb_rate: float) -> Dict[str, float]:
    rates: Dict[str, float] = {}
    for key in ("usd_to_krw_rate", "usd_to_krw"):
        value = safe_float(data.get(key))
        if value is not None and value > 1:
            rates["usd_to_krw_rate"] = value
            break
    krw_to_usd = safe_float(data.get("krw_to_usd_rate")) or safe_float(data.get("krw_to_usd"))
    if "usd_to_krw_rate" not in rates and krw_to_usd and 0 < krw_to_usd < 1:
        rates["usd_to_krw_rate"] = 1 / krw_to_usd
    for key in ("krw_to_rmb_rate", "krw_to_rmb", "rate", "exchange_rate"):
        value = safe_float(data.get(key))
        if value is not None and 0 < value < 1:
            rates["krw_to_rmb_rate"] = value
            break
    usd_to_krw = rates.get("usd_to_krw_rate")
    if "krw_to_rmb_rate" not in rates and usd_to_krw and usd_to_rmb_rate:
        rates["krw_to_rmb_rate"] = usd_to_rmb_rate / usd_to_krw
    return rates


def format_exchange_rate(rate: float) -> str:
    return f"{rate:.10f}".rstrip("0").rstrip(".")


def call_openai_exchange_rate_extraction(
    image_paths: Sequence[Path],
    usd_to_rmb_rate: float,
    progress: Optional[Callable[[str], None]] = None,
) -> Dict[str, float]:
    api_key = get_openai_api_key()
    if not api_key:
        raise RuntimeError("No OpenAI API key was found, so the exchange rate image could not be read.")
    existing_paths = [path for path in image_paths if path.exists()]
    if not existing_paths:
        raise RuntimeError("No readable exchange rate images were found.")

    def notify(message: str) -> None:
        if progress:
            try:
                progress(message)
            except Exception:
                pass

    notify("Reading exchange rate image")
    prompt = f"""
Read these exchange-rate screenshots for the Korea reimbursement form.
Return only compact JSON with these keys:
usd_to_krw_rate, krw_to_usd_rate, krw_to_rmb_rate, confidence_notes.

Rules:
- Prefer an explicit USD -> KRW value, for example 1548.86, when a screenshot shows
  1 USD = 1548.86 KRW.
- If a screenshot shows KRW -> USD instead, return krw_to_usd_rate and leave
  usd_to_krw_rate empty.
- Prefer an explicit KRW -> RMB or 汇率 value, for example 0.0044029590.
- If the images only show USD -> KRW, for example 1 USD = 1548.86 KRW,
  calculate KRW -> RMB as USD_TO_RMB / USD_TO_KRW.
- The current USD_TO_RMB value from the app is {usd_to_rmb_rate:g}.
- Return numbers as plain decimals without currency symbols or commas.
- If unsure, leave krw_to_rmb_rate empty and explain briefly in confidence_notes.
""".strip()
    content: List[Dict[str, Any]] = [{"type": "text", "text": prompt}]
    for path in existing_paths:
        content.append({"type": "image_url", "image_url": {"url": image_data_url(path)}})
    payload = json.dumps(
        {
            "model": os.getenv("OPENAI_MODEL", DEFAULT_OPENAI_MODEL),
            "messages": [
                {
                    "role": "system",
                    "content": "You extract numeric exchange rates from screenshots for reimbursement forms.",
                },
                {"role": "user", "content": content},
            ],
            "temperature": 0,
        }
    ).encode("utf-8")
    req = request.Request(
        OPENAI_CHAT_COMPLETIONS_URL,
        data=payload,
        method="POST",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
    )
    notify("Contacting AI")
    try:
        with request.urlopen(req, timeout=90) as response:
            body = response.read().decode("utf-8")
    except error.HTTPError as exc:
        try:
            details = exc.read().decode("utf-8")
        except Exception:
            details = str(exc)
        raise RuntimeError(f"OpenAI API error ({exc.code}): {details}") from exc
    except Exception as exc:
        raise RuntimeError(f"Could not contact OpenAI: {exc}") from exc

    try:
        parsed = json.loads(body)
        response_text = parsed["choices"][0]["message"]["content"]
    except Exception as exc:
        raise RuntimeError("OpenAI returned an unexpected exchange-rate response format.") from exc
    data = extract_json_object(str(response_text))
    rates = exchange_rates_from_data(data, usd_to_rmb_rate)
    if not rates:
        raise RuntimeError("AI could not find a usable exchange rate.")
    return rates


def ensure_dependencies() -> None:
    if tk is None or ttk is None:
        raise RuntimeError("Tkinter is required to run the desktop app.")
    if load_workbook is None or Workbook is None or XLImage is None:
        raise RuntimeError("openpyxl is required. Install dependencies from requirements.txt.")
    if Image is None or ImageTk is None:
        raise RuntimeError("Pillow is required. Install dependencies from requirements.txt.")


def safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def normalize_currency(value: Any, default: str = "USD") -> str:
    text = str(value or "").strip()
    if not text:
        return default
    upper = text.upper()
    if "$" in upper or "USD" in upper or "US DOLLAR" in upper or "DOLLAR" in upper:
        return "USD"
    if "KRW" in upper or "WON" in upper or "\u20a9" in upper or "\uffe6" in upper:
        return "KRW"
    if "RMB" in upper or "CNY" in upper or "YUAN" in upper or "CN\u00a5" in upper:
        return "RMB"
    if upper in {"USD", "KRW", "RMB", "CNY"}:
        return upper
    if upper in {"\u00a5", "\uffe5"}:
        return "RMB"
    return default


def category_value_to_key(value: str, form_version: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return "other"
    if " - " in raw:
        raw = raw.split(" - ", 1)[0].strip()
    normalized = normalized_category(raw, form_version)
    if normalized != "other" or raw.lower() == "other":
        return normalized
    for labels in CATEGORY_LABELS_BY_LANGUAGE.values():
        for key, label in labels.items():
            if raw == label:
                return normalized_category(key, form_version)
    return normalized


def format_amount(value: float) -> str:
    if abs(value) >= 1000:
        return f"{value:.2f}"
    return f"{value:.2f}"


def truncate_amount(value: Optional[float]) -> Optional[int]:
    if value is None:
        return None
    return math.trunc(value)


def parse_date_value(value: str) -> Any:
    text = (value or "").strip()
    if not text:
        return None
    formats = [
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y.%m.%d",
        "%Y/%m/%d",
        "%m-%d-%Y",
        "%B %d, %Y",
        "%b %d, %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return text


def excel_date_formula_or_value(value: str) -> Any:
    parsed = parse_date_value(value)
    if isinstance(parsed, date):
        return parsed
    return parsed


def normalized_category(category: str, form_version: str) -> str:
    raw = (category or "").strip().lower()
    if is_korea_other_text(raw):
        return "other"
    lookup = {
        "transport": "transportation",
        "transportation": "transportation",
        "parking": "transportation",
        "gas": "transportation",
        "fuel": "transportation",
        "uber": "transportation",
        "taxi": "transportation",
        "lodging": "lodging",
        "hotel": "lodging",
        "baggage": "lodging",
        "meal": "meals",
        "meals": "meals",
        "food": "meals",
        "restaurant": "meals",
        "entertainment": "entertainment",
        "advertising": "advertising",
        "office": "office",
        "material": "materials" if form_version == "Korea" else "office",
        "materials": "materials" if form_version == "Korea" else "office",
        "consumable": "consumables" if form_version == "Korea" else "office",
        "consumables": "consumables" if form_version == "Korea" else "office",
        "courier": "courier",
        "physical_exam": "physical_exam",
        "physical exam": "physical_exam",
        "nucleic_test": "nucleic_test",
        "nucleic test": "nucleic_test",
        "welfare": "welfare",
        "other": "other",
    }
    if raw in lookup:
        return lookup[raw]
    if form_version == "Korea" and raw in KOREA_CATEGORY_COLUMNS:
        return raw
    if form_version == "USA" and raw in USA_CATEGORY_ROWS:
        return raw
    return "other"


def is_korea_other_text(value: str) -> bool:
    raw = (value or "").lower()
    return bool(re.search(r"\b(e-?sim|sim card|data plan|internet access)\b", raw)) or any(
        token in raw for token in ("\u6d41\u91cf", "\u865a\u62df\u5361", "\uc720\uc2ec", "\ub370\uc774\ud130")
    )


def report_category_for_item(item: Any, form_version: str) -> str:
    version = normalized_form_version(form_version)
    category = category_value_to_key(str(getattr(item, "category", "") or ""), version)
    if version == "Korea":
        searchable = " ".join(
            str(getattr(item, attr, "") or "")
            for attr in ("category", "purpose", "details", "receipt_label", "place", "filename")
        )
        if is_korea_other_text(searchable):
            return "other"
        if category in KOREA_CATEGORY_COLUMNS:
            return category
        if category in {"advertising", "office"}:
            return "materials"
        if category == "entertainment":
            return "meals"
        return "other"
    if category in USA_CATEGORY_ROWS:
        return category
    if category in {"materials", "consumables"}:
        return "office"
    return "other"


def date_sort_value(value: str) -> float:
    parsed = parse_date_value(value)
    if isinstance(parsed, date):
        return float(parsed.toordinal())
    return float("inf")


def sort_receipts_for_report(items: List[Any], form_version: str) -> List[Any]:
    version = normalized_form_version(form_version)
    order = KOREA_REPORT_CATEGORY_ORDER if version == "Korea" else USA_REPORT_CATEGORY_ORDER
    order_index = {category: index for index, category in enumerate(order)}
    indexed = [
        (
            order_index.get(report_category_for_item(item, version), len(order)),
            date_sort_value(str(getattr(item, "date", "") or "")),
            index,
            item,
        )
        for index, item in enumerate(items)
    ]
    return [entry[3] for entry in sorted(indexed, key=lambda entry: (entry[0], entry[1], entry[2]))]


def normalize_rotation(value: Any) -> int:
    try:
        degrees = int(float(value))
    except (TypeError, ValueError):
        degrees = 0
    return degrees % 360 // 90 * 90


def oriented_image_from_path(path: Path, rotation_degrees: Any = 0) -> Any:
    with Image.open(path) as source:
        image = ImageOps.exif_transpose(source) if ImageOps is not None else source.copy()
    rotation = normalize_rotation(rotation_degrees)
    if rotation:
        image = image.rotate(-rotation, expand=True)
    return image


def default_crop_points(width: int, height: int) -> List[Tuple[float, float]]:
    return [
        (0.0, 0.0),
        (float(width), 0.0),
        (float(width), float(height)),
        (0.0, float(height)),
    ]


def normalized_crop_points(crop_box: Any, width: int, height: int) -> Optional[List[Tuple[float, float]]]:
    if width <= 0 or height <= 0:
        return None
    if not isinstance(crop_box, (list, tuple)) or len(crop_box) != 4:
        if not isinstance(crop_box, (list, tuple)) or len(crop_box) != 8:
            return None
    try:
        values = [float(value) for value in crop_box]
    except (TypeError, ValueError):
        return None
    if len(values) == 4:
        x1, y1, x2, y2 = values
        points = [(x1, y1), (x2, y1), (x2, y2), (x1, y2)]
    else:
        points = [
            (values[0], values[1]),
            (values[2], values[3]),
            (values[4], values[5]),
            (values[6], values[7]),
        ]
    clamped = [
        (max(0.0, min(float(width), x)), max(0.0, min(float(height), y)))
        for x, y in points
    ]
    top_width = math.dist(clamped[0], clamped[1])
    bottom_width = math.dist(clamped[3], clamped[2])
    left_height = math.dist(clamped[0], clamped[3])
    right_height = math.dist(clamped[1], clamped[2])
    if max(top_width, bottom_width) < 5 or max(left_height, right_height) < 5:
        return None
    default = default_crop_points(width, height)
    if all(math.dist(point, default_point) < 1.0 for point, default_point in zip(clamped, default)):
        return None
    return clamped


def normalized_crop_box(crop_box: Any, width: int, height: int) -> Optional[Tuple[int, int, int, int]]:
    points = normalized_crop_points(crop_box, width, height)
    if not points:
        return None
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    return (
        int(round(max(0.0, min(xs)))),
        int(round(max(0.0, min(ys)))),
        int(round(min(float(width), max(xs)))),
        int(round(min(float(height), max(ys)))),
    )


def solve_linear_system(matrix: List[List[float]], vector: List[float]) -> List[float]:
    size = len(vector)
    rows = [matrix[index][:] + [vector[index]] for index in range(size)]
    for pivot_index in range(size):
        pivot_row = max(range(pivot_index, size), key=lambda row: abs(rows[row][pivot_index]))
        if abs(rows[pivot_row][pivot_index]) < 1e-9:
            raise ValueError("Perspective crop points are too close together.")
        rows[pivot_index], rows[pivot_row] = rows[pivot_row], rows[pivot_index]
        pivot = rows[pivot_index][pivot_index]
        rows[pivot_index] = [value / pivot for value in rows[pivot_index]]
        for row_index in range(size):
            if row_index == pivot_index:
                continue
            factor = rows[row_index][pivot_index]
            rows[row_index] = [
                value - factor * rows[pivot_index][col_index]
                for col_index, value in enumerate(rows[row_index])
            ]
    return [rows[index][-1] for index in range(size)]


def perspective_coefficients(
    source_points: List[Tuple[float, float]],
    width: int,
    height: int,
) -> List[float]:
    destination_points = [
        (0.0, 0.0),
        (float(width), 0.0),
        (float(width), float(height)),
        (0.0, float(height)),
    ]
    matrix: List[List[float]] = []
    vector: List[float] = []
    for (u, v), (x, y) in zip(destination_points, source_points):
        matrix.append([u, v, 1.0, 0.0, 0.0, 0.0, -u * x, -v * x])
        vector.append(x)
        matrix.append([0.0, 0.0, 0.0, u, v, 1.0, -u * y, -v * y])
        vector.append(y)
    return solve_linear_system(matrix, vector)


def perspective_crop_image(image: Any, crop_box: Any) -> Any:
    points = normalized_crop_points(crop_box, image.width, image.height)
    if not points:
        return image
    top_width = math.dist(points[0], points[1])
    bottom_width = math.dist(points[3], points[2])
    left_height = math.dist(points[0], points[3])
    right_height = math.dist(points[1], points[2])
    output_width = max(1, int(round(max(top_width, bottom_width))))
    output_height = max(1, int(round(max(left_height, right_height))))
    try:
        coefficients = perspective_coefficients(points, output_width, output_height)
    except Exception:
        setup_logging()
        LOGGER.exception("Could not solve perspective crop; using original image")
        return image
    transform_mode = getattr(Image, "Transform", Image).PERSPECTIVE
    resample = getattr(getattr(Image, "Resampling", Image), "BICUBIC", Image.BICUBIC)
    return image.transform((output_width, output_height), transform_mode, coefficients, resample)


def flatten_crop_points(points: List[Tuple[float, float]]) -> List[float]:
    return [round(value, 2) for point in points for value in point]


def rotated_crop_points(
    crop_box: Any,
    width: int,
    height: int,
    delta_degrees: int,
) -> Optional[List[float]]:
    points = normalized_crop_points(crop_box, width, height)
    if not points:
        return None
    delta = normalize_rotation(delta_degrees)
    if delta == 90:
        transformed = [(height - y, x) for x, y in points]
        ordered = [transformed[index] for index in (3, 0, 1, 2)]
    elif delta == 270:
        transformed = [(y, width - x) for x, y in points]
        ordered = [transformed[index] for index in (1, 2, 3, 0)]
    elif delta == 180:
        transformed = [(width - x, height - y) for x, y in points]
        ordered = [transformed[index] for index in (2, 3, 0, 1)]
    else:
        ordered = points
    return flatten_crop_points(ordered)


def prepare_receipt_image_file(path: Path, crop_box: Any = None, rotation_degrees: Any = 0) -> Path:
    if Image is None:
        return path
    target_dir = WORK_DIR / "prepared_images"
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"{uuid.uuid4().hex}.png"
    image = oriented_image_from_path(path, rotation_degrees)
    image = perspective_crop_image(image, crop_box)
    if image.mode not in {"RGB", "RGBA"}:
        image = image.convert("RGB")
    image.save(target, "PNG")
    return target


def prepare_excel_image_file(path: Path, crop_box: Any = None, rotation_degrees: Any = 0) -> Path:
    return prepare_receipt_image_file(path, crop_box, rotation_degrees)


def resize_excel_image(
    path: Path,
    max_width: int,
    max_height: int,
    crop_box: Any = None,
    rotation_degrees: Any = 0,
) -> Any:
    excel_path = prepare_excel_image_file(path, crop_box, rotation_degrees)
    img = XLImage(str(excel_path))
    if Image is None:
        img.width = max_width
        img.height = max_height
        return img
    with Image.open(excel_path) as pil_img:
        width, height = pil_img.size
    if width <= 0 or height <= 0:
        img.width = max_width
        img.height = max_height
        return img
    scale = min(max_width / width, max_height / height, 1.0)
    img.width = max(1, int(width * scale))
    img.height = max(1, int(height * scale))
    return img


def ensure_receipt_images(item: ReceiptItem) -> List[Dict[str, Any]]:
    normalized = [attachment for attachment in (normalize_attachment(raw) for raw in item.receipt_images) if attachment]
    if not normalized and item.path:
        normalized = [attachment_from_item(item)]
    item.receipt_images = normalized
    if normalized:
        primary = normalized[0]
        item.path = primary["path"]
        item.filename = primary["filename"]
        item.source_path = primary.get("source_path", "")
        item.source_page = primary.get("source_page", "")
        item.crop_box = primary.get("crop_box")
        item.rotation_degrees = normalize_rotation(primary.get("rotation_degrees", 0))
    return item.receipt_images


def attachment_file_key(attachment: Dict[str, Any]) -> Tuple[str, str]:
    return selected_file_key(
        Path(str(attachment.get("path") or "")),
        str(attachment.get("source_path") or ""),
        str(attachment.get("source_page") or ""),
    )


def merge_attachment_lists(
    target: List[Dict[str, Any]],
    source: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    seen = {attachment_file_key(attachment) for attachment in target}
    for attachment in source:
        key = attachment_file_key(attachment)
        if key in seen:
            continue
        target.append(copy.deepcopy(attachment))
        seen.add(key)
    return target


def prepare_attachment_image_file(attachment: Dict[str, Any]) -> Path:
    return prepare_receipt_image_file(
        Path(str(attachment.get("path") or "")),
        attachment.get("crop_box"),
        attachment.get("rotation_degrees", 0),
    )


def prepare_attachment_contact_sheet(
    attachments: List[Dict[str, Any]],
    max_width: int,
    max_height: int,
    allow_upscale: bool = False,
) -> Optional[Path]:
    if Image is None:
        return None
    normalized = [attachment for attachment in (normalize_attachment(raw) for raw in attachments) if attachment]
    if not normalized:
        return None
    if len(normalized) == 1 and not allow_upscale:
        return prepare_attachment_image_file(normalized[0])
    target_dir = WORK_DIR / "prepared_images"
    target_dir.mkdir(parents=True, exist_ok=True)
    prepared_images = []
    for attachment in normalized:
        path = Path(str(attachment.get("path") or ""))
        if not path.exists():
            continue
        prepared = prepare_attachment_image_file(attachment)
        with Image.open(prepared) as img:
            prepared_images.append(img.convert("RGB").copy())
    if not prepared_images:
        return None
    cols = min(2, len(prepared_images))
    rows = (len(prepared_images) + cols - 1) // cols
    gutter = 8
    cell_width = max(1, (max_width - gutter * (cols - 1)) // cols)
    cell_height = max(1, (max_height - gutter * (rows - 1)) // rows)
    sheet = Image.new("RGB", (max_width, max_height), "white")
    for index, img in enumerate(prepared_images):
        tile = img.copy()
        if allow_upscale:
            scale = min(cell_width / max(1, tile.width), cell_height / max(1, tile.height))
            resampling = getattr(Image, "Resampling", None)
            resample = getattr(resampling, "LANCZOS", getattr(Image, "LANCZOS", Image.BICUBIC))
            tile = tile.resize((max(1, int(tile.width * scale)), max(1, int(tile.height * scale))), resample)
        else:
            tile.thumbnail((cell_width, cell_height))
        col = index % cols
        row = index // cols
        x = col * (cell_width + gutter) + max(0, (cell_width - tile.width) // 2)
        y = row * (cell_height + gutter) + max(0, (cell_height - tile.height) // 2)
        sheet.paste(tile, (x, y))
    target = target_dir / f"{uuid.uuid4().hex}_contact.png"
    sheet.save(target, "PNG")
    return target


def resize_attachment_contact_sheet(
    attachments: List[Dict[str, Any]],
    max_width: int,
    max_height: int,
    allow_upscale: bool = False,
) -> Optional[Any]:
    contact_path = prepare_attachment_contact_sheet(attachments, max_width, max_height, allow_upscale)
    if contact_path is None:
        return None
    img = XLImage(str(contact_path))
    with Image.open(contact_path) as pil_img:
        width, height = pil_img.size
    scale = min(max_width / max(1, width), max_height / max(1, height))
    if not allow_upscale:
        scale = min(scale, 1.0)
    img.width = max(1, int(width * scale))
    img.height = max(1, int(height * scale))
    return img


def clear_images(sheet: Any) -> None:
    try:
        sheet._images = []
    except Exception:
        pass


def clear_cells(sheet: Any, rows: Iterable[int], cols: Sequence[str]) -> None:
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            if MergedCell is not None and isinstance(cell, MergedCell):
                continue
            cell.value = None


def fit_columns_for_receipts(sheet: Any) -> None:
    for col in "ABCDEFGH":
        sheet.column_dimensions[col].width = 14


def configure_korea_receipt_page(sheet: Any, slot_count: int) -> None:
    page_height = 60
    slots_per_page = 4
    page_count = max(1, (slot_count + slots_per_page - 1) // slots_per_page)
    last_row = page_count * page_height
    sheet.print_area = f"A1:H{last_row}"
    if sheet.sheet_properties.pageSetUpPr is None:
        try:
            from openpyxl.worksheet.properties import PageSetupProperties

            sheet.sheet_properties.pageSetUpPr = PageSetupProperties()
        except Exception:
            pass
    if sheet.sheet_properties.pageSetUpPr is not None:
        sheet.sheet_properties.pageSetUpPr.fitToPage = False
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.scale = 78
    sheet.page_setup.fitToWidth = None
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.45
    sheet.page_margins.right = 0.45
    sheet.page_margins.top = 0.45
    sheet.page_margins.bottom = 0.45
    try:
        from openpyxl.worksheet.pagebreak import Break

        sheet.row_breaks.brk = []
        for row in range(page_height, last_row, page_height):
            sheet.row_breaks.append(Break(id=row))
    except Exception:
        pass


def korea_receipt_slot(index: int, wide: bool = False) -> Dict[str, Any]:
    page_height = 60
    slots_per_page = 4
    row_offsets = [1, 1, 31, 31]
    col_ranges = [("A", "D"), ("E", "H"), ("A", "D"), ("E", "H")]
    page = index // slots_per_page
    slot = index % slots_per_page
    label_row = page * page_height + row_offsets[slot]
    image_row = label_row + 1
    if wide:
        return {
            "label_range": f"A{label_row}:H{label_row}",
            "label_cell": f"A{label_row}",
            "image_cell": f"A{image_row}",
            "max_width": 760,
            "max_height": 520,
        }
    start_col, end_col = col_ranges[slot]
    return {
        "label_range": f"{start_col}{label_row}:{end_col}{label_row}",
        "label_cell": f"{start_col}{label_row}",
        "image_cell": f"{start_col}{image_row}",
        "max_width": 370,
        "max_height": 520,
    }


def korea_receipt_cost_text(item: "ReceiptItem") -> str:
    amount = item.amount.strip()
    if amount:
        return f"{amount} {normalize_currency(item.currency, 'KRW')}".strip()
    if item.krw_amount.strip():
        return f"{item.krw_amount.strip()} KRW"
    if item.rmb_amount.strip():
        return f"{item.rmb_amount.strip()} RMB"
    return ""


def korea_receipt_payment_label(index: int, item: "ReceiptItem") -> str:
    details: List[str] = []
    if item.date.strip():
        details.append(item.date.strip())
    content = (
        item.purpose.strip()
        or item.details.strip()
        or item.receipt_label.strip()
        or item.place.strip()
        or item.filename
    )
    if content:
        details.append(content)
    cost = korea_receipt_cost_text(item)
    if cost:
        details.append(cost)
    return " | ".join(details) if details else item.filename or f"Payment {index}"


def set_korea_receipt_label(sheet: Any, cell_ref: str, label: str) -> None:
    cell = sheet[cell_ref]
    cell.value = label
    if Font is not None:
        cell.font = Font(bold=True, size=10, color="1F2937")
    if Alignment is not None:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if PatternFill is not None:
        cell.fill = PatternFill(fill_type="solid", fgColor="E9EEF3")


def add_korea_receipt_block(
    sheet: Any,
    index: int,
    label: str,
    attachments: List[Dict[str, Any]],
    wide: bool = False,
) -> None:
    slot = korea_receipt_slot(index, wide=wide)
    label_row = int(re.sub(r"^[A-Z]+", "", str(slot["label_cell"])))
    try:
        sheet.merge_cells(str(slot["label_range"]))
    except Exception:
        pass
    sheet.row_dimensions[label_row].height = 22
    set_korea_receipt_label(sheet, str(slot["label_cell"]), label)
    receipt_image = resize_attachment_contact_sheet(
        attachments,
        int(slot["max_width"]),
        int(slot["max_height"]),
        allow_upscale=True,
    )
    if receipt_image is not None:
        receipt_image.width = int(slot["max_width"])
        receipt_image.height = int(slot["max_height"])
        sheet.add_image(receipt_image, str(slot["image_cell"]))


def export_usa(
    items: List[ReceiptItem],
    output_path: Path,
    exchange_rate: float,
    bank_items: Optional[List[BankStatementItem]] = None,
) -> None:
    template = TEMPLATE_DIR / USA_TEMPLATE_NAME
    if not template.exists():
        raise FileNotFoundError(f"Missing USA template: {template}")
    wb = load_workbook(template)
    ws = wb["Expense report"]
    receipts_ws = wb["Receipt and Payment of expenses"]
    sorted_items = sort_receipts_for_report(items, "USA")

    ws["A3"] = f"Date / 填表日期： {date.today().strftime('%m/%d/%Y')}"
    ws["A4"] = "Employee: / 申请人："
    ws["J1"] = exchange_rate
    all_entry_rows = sorted({row for rows in USA_CATEGORY_ROWS.values() for row in rows})
    clear_cells(ws, all_entry_rows, ["A", "B", "C", "D", "E", "F", "H"])
    for row in all_entry_rows:
        ws[f"G{row}"] = f"=F{row}*$J$1"

    row_cursors = {cat: 0 for cat in USA_CATEGORY_ROWS}
    for item in sorted_items:
        category = report_category_for_item(item, "USA")
        if category not in USA_CATEGORY_ROWS:
            category = "other"
        rows = USA_CATEGORY_ROWS[category]
        cursor = row_cursors[category]
        if cursor >= len(rows):
            raise RuntimeError(f"Not enough USA template rows for {USA_CATEGORY_LABELS.get(category, category)}.")
        row = rows[cursor]
        row_cursors[category] += 1
        amount = safe_float(item.amount)
        ws[f"A{row}"] = item.place.strip()
        ws[f"B{row}"] = excel_date_formula_or_value(item.date)
        ws[f"C{row}"] = item.details.strip() or item.receipt_label.strip() or item.filename
        ws[f"D{row}"] = item.purpose.strip()
        ws[f"E{row}"] = item.project_number.strip()
        ws[f"F{row}"] = amount if amount is not None else None
        ws[f"G{row}"] = f"=F{row}*$J$1"

    clear_images(receipts_ws)
    for row in range(2, max(56, len(sorted_items) + 3)):
        for col in "ABCDE":
            receipts_ws[f"{col}{row}"] = None
    receipts_ws.column_dimensions["D"].width = 38
    receipts_ws.column_dimensions["E"].width = 26
    proof_by_receipt_id: Dict[str, List[Dict[str, Any]]] = {}
    for bank_item in bank_items or []:
        if bank_item.matched_receipt_id:
            proof_by_receipt_id.setdefault(bank_item.matched_receipt_id, []).append(attachment_from_item(bank_item))
    for index, item in enumerate(sorted_items, start=1):
        row = index + 1
        receipts_ws[f"A{row}"] = index
        receipts_ws[f"B{row}"] = excel_date_formula_or_value(item.date)
        receipts_ws[f"C{row}"] = safe_float(item.amount)
        receipts_ws.row_dimensions[row].height = 126
        receipt_image = resize_attachment_contact_sheet(ensure_receipt_images(item), 260, 150)
        if receipt_image is not None:
            receipts_ws.add_image(receipt_image, f"D{row}")
        proof_image = resize_attachment_contact_sheet(proof_by_receipt_id.get(item.item_id, []), 180, 150)
        if proof_image is not None:
            receipts_ws.add_image(proof_image, f"E{row}")

    wb.save(output_path)


def clone_sheet(
    source_ws: Any,
    target_wb: Any,
    title: str,
    index: Optional[int] = None,
    min_rows: int = 0,
    min_cols: int = 0,
) -> Any:
    target_ws = target_wb.create_sheet(title=title, index=index)
    max_row = max(source_ws.max_row or 1, min_rows)
    max_col = max(source_ws.max_column or 1, min_cols)
    for row in source_ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for source_cell in row:
            target_cell = target_ws[source_cell.coordinate]
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy.copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.font:
                target_cell.font = copy.copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy.copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy.copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy.copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy.copy(source_cell.protection)
            if source_cell.comment:
                target_cell.comment = copy.copy(source_cell.comment)
    for merged in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged))
    for key, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key] = copy.copy(dim)
    for key, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[key] = copy.copy(dim)
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.sheet_format = copy.copy(source_ws.sheet_format)
    target_ws.sheet_properties = copy.copy(source_ws.sheet_properties)
    target_ws.page_margins = copy.copy(source_ws.page_margins)
    target_ws.page_setup = copy.copy(source_ws.page_setup)
    target_ws.print_options = copy.copy(source_ws.print_options)
    if source_ws.print_area:
        target_ws.print_area = source_ws.print_area
    return target_ws


def clear_freeze_panes(sheet: Any) -> None:
    sheet.freeze_panes = None
    sheet.sheet_view.pane = None
    if Selection is not None:
        sheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]


def korea_amounts(
    item: ReceiptItem,
    krw_to_rmb_rate: float,
    usd_to_rmb_rate: float,
    usd_to_krw_rate: float,
) -> Tuple[Optional[float], Optional[float], str]:
    amount = safe_float(item.amount)
    krw = safe_float(item.krw_amount)
    rmb = safe_float(item.rmb_amount)
    currency = normalize_currency(item.currency, "KRW")
    if amount is not None:
        if currency == "USD":
            krw = amount * usd_to_krw_rate if usd_to_krw_rate else krw
            if krw is None and usd_to_rmb_rate and krw_to_rmb_rate:
                krw = (amount * usd_to_rmb_rate) / krw_to_rmb_rate
            rmb = krw * krw_to_rmb_rate if krw is not None and krw_to_rmb_rate else (amount * usd_to_rmb_rate if usd_to_rmb_rate else rmb)
        elif currency == "RMB" or currency == "CNY":
            rmb = amount
            krw = amount / krw_to_rmb_rate if krw_to_rmb_rate else (
                (amount / usd_to_rmb_rate) * usd_to_krw_rate if usd_to_rmb_rate and usd_to_krw_rate else None
            )
        elif currency == "KRW":
            krw = amount
            rmb = amount * krw_to_rmb_rate if krw_to_rmb_rate else None
    if krw is None and rmb is not None and krw_to_rmb_rate:
        krw = rmb / krw_to_rmb_rate
    if rmb is None and krw is not None and krw_to_rmb_rate:
        rmb = krw * krw_to_rmb_rate
    note = ""
    if amount is not None and currency not in {"KRW", "RMB", "CNY"}:
        note = f"({amount:g} {currency})"
    return krw, rmb, note


def korea_cover_bucket(category: str) -> str:
    if category == "transportation":
        return "transportation"
    if category == "lodging":
        return "lodging"
    if category in {"meals", "entertainment"}:
        return "meals"
    if category in {"materials", "consumables", "office"}:
        return "consumables"
    return "other"


def export_korea(
    items: List[ReceiptItem],
    output_path: Path,
    krw_to_rmb_rate: float,
    usd_to_rmb_rate: float,
    usd_to_krw_rate: float,
    exchange_rate_items: Optional[List[BankStatementItem]] = None,
) -> None:
    cover_template = TEMPLATE_DIR / KOREA_COVER_TEMPLATE_NAME
    details_template = TEMPLATE_DIR / KOREA_DETAILS_TEMPLATE_NAME
    if not cover_template.exists():
        raise FileNotFoundError(f"Missing Korea cover template: {cover_template}")
    if not details_template.exists():
        raise FileNotFoundError(f"Missing Korea details template: {details_template}")

    wb = load_workbook(cover_template)
    detail_template_wb = load_workbook(details_template)
    cover_ws = wb.worksheets[0]
    receipts_ws = wb.worksheets[1]
    detail_ws = clone_sheet(detail_template_wb.worksheets[0], wb, "报销明细", index=0, min_rows=35, min_cols=20)
    sorted_items = sort_receipts_for_report(items, "Korea")
    try:
        wb._sheets = [detail_ws, cover_ws, receipts_ws]
    except Exception:
        pass
    clear_freeze_panes(detail_ws)
    detail_ws.print_area = "A1:S35"
    if detail_ws.sheet_properties.pageSetUpPr is not None:
        detail_ws.sheet_properties.pageSetUpPr.fitToPage = False
    detail_ws.page_setup.orientation = "landscape"
    detail_ws.page_setup.scale = 34
    detail_ws.page_setup.fitToWidth = None
    detail_ws.page_setup.fitToHeight = None
    detail_ws.page_margins.left = 0.25
    detail_ws.page_margins.right = 0.25
    detail_ws.page_margins.top = 0.75
    detail_ws.page_margins.bottom = 0.75
    for col in "KLMNOPQRS":
        detail_ws.column_dimensions[col].hidden = False

    cover_ws["A2"] = f"报销部门：  {date.today().year}年 {date.today().month}月 {date.today().day}日 填 单据及附件共  页"
    cover_ws["A11"] = "领导审批           会计主管              会计                  出纳                 报销人                   领款人 "
    for row, label in [(row, label) for row, label in KOREA_COVER_ROWS.values()]:
        cover_ws[f"A{row}"] = label
        cover_ws[f"C{row}"] = 0
        cover_ws[f"D{row}"] = 0
    cover_ws["C9"] = "=SUM(C4:C8)"
    cover_ws["D9"] = "=SUM(D4:D8)"
    cover_ws["B10"] = "=D9"

    for merged_range in ("A33:B33", "A34:B34", "A35:B35"):
        try:
            detail_ws.unmerge_cells(merged_range)
        except Exception:
            pass
    clear_cells(detail_ws, range(3, 36), list("ABCDEFGHIJKLMNOPQRS"))
    try:
        detail_ws.merge_cells("A33:B33")
        detail_ws.merge_cells("A34:B34")
    except Exception:
        pass
    detail_ws["A33"] = "合计（外币）\nTotal"
    detail_ws["Q33"] = "=SUM(Q3:Q32)"
    detail_ws["Q33"].number_format = KRW_NUMBER_FORMAT
    detail_ws["A34"] = "合计（人民币）\nTotal"
    detail_ws["R34"] = "=SUM(R3:R33)"

    summary: Dict[str, Tuple[float, float]] = {
        key: (0.0, 0.0) for key in KOREA_COVER_ROWS
    }
    for index, item in enumerate(sorted_items, start=3):
        if index > 32:
            raise RuntimeError("Korea template supports up to 30 detail rows in this version.")
        category = report_category_for_item(item, "Korea")
        if category not in KOREA_CATEGORY_COLUMNS:
            category = "other"
        krw, rmb, original_note = korea_amounts(item, krw_to_rmb_rate, usd_to_rmb_rate, usd_to_krw_rate)
        krw_whole = truncate_amount(krw)
        detail_ws[f"A{index}"] = excel_date_formula_or_value(item.date)
        detail_ws[f"B{index}"] = item.purpose.strip() or item.details.strip()
        detail_ws[f"C{index}"] = item.place.strip()
        detail_ws[f"D{index}"] = ""
        detail_ws[f"E{index}"] = item.project_number.strip()
        category_col = KOREA_CATEGORY_COLUMNS[category]
        detail_ws[f"{category_col}{index}"] = krw_whole
        detail_ws[f"{category_col}{index}"].number_format = KRW_NUMBER_FORMAT
        detail_ws[f"P{index}"] = original_note
        detail_ws[f"Q{index}"] = krw_whole
        detail_ws[f"Q{index}"].number_format = KRW_NUMBER_FORMAT
        detail_ws[f"R{index}"] = rmb
        detail_ws[f"S{index}"] = item.payment_method.strip()
        bucket = korea_cover_bucket(category)
        cur_krw, cur_rmb = summary[bucket]
        summary[bucket] = (cur_krw + (krw_whole or 0), cur_rmb + (rmb or 0.0))

    for bucket, (krw_total, rmb_total) in summary.items():
        row, _label = KOREA_COVER_ROWS[bucket]
        cover_ws[f"C{row}"] = truncate_amount(krw_total) if krw_total else None
        cover_ws[f"C{row}"].number_format = KRW_NUMBER_FORMAT
        cover_ws[f"D{row}"] = round(rmb_total, 2) if rmb_total else None
    cover_ws["C9"].number_format = KRW_NUMBER_FORMAT

    clear_images(receipts_ws)
    fit_columns_for_receipts(receipts_ws)
    exchange_rate_attachments = [attachment_from_item(item) for item in exchange_rate_items or []]
    slot_count = len(sorted_items) + (2 if exchange_rate_attachments else 0)
    configure_korea_receipt_page(receipts_ws, slot_count)
    page_height = 60
    slots_per_page = 4
    page_count = max(1, (slot_count + slots_per_page - 1) // slots_per_page)
    last_receipt_row = page_count * page_height
    for row in range(1, max(240, last_receipt_row + 1)):
        for col in "ABCDEFGH":
            receipts_ws[f"{col}{row}"] = None
    block_index = 0
    if exchange_rate_attachments:
        add_korea_receipt_block(receipts_ws, block_index, "汇率 / Exchange Rate", exchange_rate_attachments, wide=True)
        block_index += 2
    for idx, item in enumerate(sorted_items):
        add_korea_receipt_block(
            receipts_ws,
            block_index,
            korea_receipt_payment_label(idx + 1, item),
            ensure_receipt_images(item),
        )
        block_index += 1

    wb.save(output_path)


class ReimbursementHelperApp:
    def __init__(self, root: Any) -> None:
        self.root = root
        self.items: List[ReceiptItem] = []
        self.bank_items: List[BankStatementItem] = []
        self.exchange_rate_items: List[BankStatementItem] = []
        self.selected_index: Optional[int] = None
        self.selected_bank_index: Optional[int] = None
        self.bank_tree: Optional[Any] = None
        self.photo: Optional[Any] = None
        self.preview_canvas: Optional[Any] = None
        self.revert_crop_btn: Optional[Any] = None
        self.rotate_left_btn: Optional[Any] = None
        self.rotate_right_btn: Optional[Any] = None
        self.delete_screenshot_btn: Optional[Any] = None
        self.swap_proof_btn: Optional[Any] = None
        self.unlink_proof_btn: Optional[Any] = None
        self.upload_folder_btn: Optional[Any] = None
        self.select_payment_proof_btn: Optional[Any] = None
        self.generate_details_btn: Optional[Any] = None
        self.generate_all_btn: Optional[Any] = None
        self.generate_excel_btn: Optional[Any] = None
        self.selected_attachment_kind = "receipt"
        self.selected_attachment_index = 0
        self._selection_anchor_index: Optional[int] = None
        self.preview_tiles: List[Dict[str, Any]] = []
        self.preview_photos: List[Any] = []
        self.preview_action_regions: Dict[str, Dict[str, Any]] = {}
        self.proof_drop_region: Optional[Tuple[int, int, int, int]] = None
        self._pressed_preview_tile: Optional[Dict[str, Any]] = None
        self._dragging_preview_tile: Optional[Dict[str, Any]] = None
        self._drag_start: Tuple[int, int] = (0, 0)
        self._drag_outline_id: Optional[int] = None
        self._preview_image_bounds: Tuple[int, int, int, int] = (0, 0, 0, 0)
        self._preview_original_size: Tuple[int, int] = (0, 0)
        self._crop_handle_centers: Dict[str, Tuple[float, float]] = {}
        self._dragging_crop_handle: Optional[str] = None
        self.settings = load_user_settings()
        self.form_version = tk.StringVar(value=normalized_form_version(self.settings.get("last_form_version", "USA")))
        self.language = tk.StringVar(value=normalize_language(self.settings.get("language", "en")))
        self.language_display = tk.StringVar(value=LANGUAGE_LABELS.get(self.language.get(), "English"))
        self.exchange_rate = tk.StringVar(value=str(self.settings.get("usa_exchange_rate", DEFAULT_USD_TO_RMB_RATE)))
        self.usd_to_krw_rate = tk.StringVar(value=str(self.settings.get("usd_to_krw_rate", DEFAULT_USD_TO_KRW_RATE)))
        self.krw_to_rmb_rate = tk.StringVar(value=str(self.settings.get("krw_to_rmb_rate", DEFAULT_KRW_TO_RMB_RATE)))
        self.status_text = tk.StringVar(value=self.tr("ready"))
        self.progress_text = tk.StringVar(value="")
        self.progress_value = tk.DoubleVar(value=0)
        self.field_vars: Dict[str, Any] = {}
        self.field_labels: Dict[str, Any] = {}
        self.field_widgets: Dict[str, Any] = {}
        self.category_values: List[str] = []
        self.category_label_to_key: Dict[str, str] = {}
        self.category_combo: Optional[Any] = None
        self.header_label: Optional[Any] = None
        self.form_label: Optional[Any] = None
        self.language_label: Optional[Any] = None
        self.manager_label: Optional[Any] = None
        self.details_label: Optional[Any] = None
        self.preview_label: Optional[Any] = None
        self.remove_btn: Optional[Any] = None
        self.clear_btn: Optional[Any] = None
        self.open_output_btn: Optional[Any] = None
        self.usa_rate_label: Optional[Any] = None
        self.usa_rate_entry: Optional[Any] = None
        self.usd_krw_rate_label: Optional[Any] = None
        self.usd_krw_rate_entry: Optional[Any] = None
        self.krw_rate_label: Optional[Any] = None
        self.krw_rate_entry: Optional[Any] = None
        self._loading_fields = False
        self._syncing_amounts = False
        self._last_amount_source = "amount"
        self._last_form_version = self.form_version.get()
        self._busy = False
        self._details_ready_for_export = False
        self.last_generated_output_path: Optional[Path] = None
        self._build_ui()
        self._attach_amount_traces()
        self._on_form_version_changed()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.after(250, self.restore_previous_session_if_available)

    def tr(self, key: str, **values: Any) -> str:
        return ui_text(self.language.get(), key, **values)

    def _build_ui(self) -> None:
        self.root.title(self.tr("app_title"))
        self.root.geometry("1220x760")
        self.root.minsize(980, 640)
        app_bg = "#E9EEF3"
        panel_bg = "#F8FAFC"
        canvas_bg = "#F5F7FA"
        border = "#CBD5E1"
        text = "#17202A"
        muted = "#64748B"
        accent = "#2563EB"
        button_bg = "#E6EDF5"
        button_active = "#D6E2EE"
        button_disabled = "#EEF3F8"
        button_border = "#94A3B8"
        control_bg = "#F1F5F9"
        self.root.configure(bg=app_bg)
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure(
            ".",
            background=app_bg,
            foreground=text,
            bordercolor=border,
            lightcolor="#F8FAFC",
            darkcolor="#AAB8C7",
            troughcolor="#D8E1EC",
            focuscolor=app_bg,
        )
        style.configure("TFrame", background=app_bg)
        style.configure("Panel.TFrame", background=panel_bg)
        style.configure("TLabel", background=app_bg, foreground=text, font=("Segoe UI", 10))
        style.configure("Header.TLabel", background=app_bg, foreground=text, font=("Segoe UI", 13, "bold"))
        style.configure("Panel.TLabel", background=panel_bg, foreground=text, font=("Segoe UI", 10))
        style.configure("PanelHeader.TLabel", background=panel_bg, foreground=text, font=("Segoe UI", 13, "bold"))
        style.configure("Muted.TLabel", background=app_bg, foreground=muted, font=("Segoe UI", 9))
        style.configure(
            "TButton",
            font=("Segoe UI", 10),
            padding=(10, 5),
            background=button_bg,
            foreground=text,
            bordercolor=button_border,
            lightcolor=button_bg,
            darkcolor=button_border,
            focuscolor=button_bg,
            relief="flat",
        )
        style.map(
            "TButton",
            background=[
                ("pressed", "#C6D7E8"),
                ("active", button_active),
                ("disabled", button_disabled),
            ],
            foreground=[("disabled", "#94A3B8")],
            bordercolor=[("disabled", "#C9D4E0"), ("active", "#7C93AA")],
            lightcolor=[("disabled", button_disabled), ("active", button_active)],
            darkcolor=[("disabled", "#C9D4E0"), ("active", "#7C93AA")],
        )
        style.configure(
            "Recommended.TButton",
            font=("Segoe UI", 10),
            padding=(10, 5),
            background=accent,
            foreground="#FFFFFF",
            bordercolor="#1D4ED8",
            lightcolor=accent,
            darkcolor="#1D4ED8",
            focuscolor=accent,
            relief="flat",
        )
        style.map(
            "Recommended.TButton",
            background=[
                ("pressed", "#1E40AF"),
                ("active", "#1D4ED8"),
                ("disabled", button_disabled),
            ],
            foreground=[("disabled", "#94A3B8"), ("!disabled", "#FFFFFF")],
            bordercolor=[("disabled", "#C9D4E0"), ("active", "#1E40AF")],
            lightcolor=[("disabled", button_disabled), ("active", "#1D4ED8")],
            darkcolor=[("disabled", "#C9D4E0"), ("active", "#1E40AF")],
        )
        style.configure(
            "Icon.TButton",
            font=("Segoe UI Symbol", 13),
            width=3,
            padding=(4, 3),
            background=button_bg,
            bordercolor=button_border,
            lightcolor=button_bg,
            darkcolor=button_border,
            focuscolor=button_bg,
            relief="flat",
        )
        style.configure(
            "TEntry",
            fieldbackground=control_bg,
            background=control_bg,
            foreground=text,
            insertcolor=text,
            bordercolor=border,
            lightcolor=control_bg,
            darkcolor=border,
            focuscolor=control_bg,
            selectbackground="#CFE1F6",
            selectforeground=text,
            relief="flat",
        )
        style.configure(
            "TCombobox",
            fieldbackground=control_bg,
            background=button_bg,
            foreground=text,
            arrowcolor=text,
            bordercolor=border,
            lightcolor=control_bg,
            darkcolor=border,
            focuscolor=control_bg,
            selectbackground="#CFE1F6",
            selectforeground=text,
            relief="flat",
        )
        style.map(
            "TCombobox",
            fieldbackground=[("readonly", control_bg), ("!disabled", control_bg), ("disabled", button_disabled)],
            background=[("readonly", button_bg), ("active", button_active), ("disabled", button_disabled)],
            selectbackground=[("readonly", control_bg), ("!disabled", control_bg), ("disabled", button_disabled)],
            selectforeground=[("readonly", text), ("!disabled", text)],
            bordercolor=[("focus", "#7EA5D3"), ("disabled", "#C9D4E0")],
            lightcolor=[("focus", control_bg), ("disabled", button_disabled)],
            darkcolor=[("focus", "#7EA5D3"), ("disabled", "#C9D4E0")],
        )
        style.configure("Treeview", background=panel_bg, fieldbackground=panel_bg, foreground=text, rowheight=24, bordercolor=border)
        style.configure(
            "Treeview.Heading",
            background="#E2EAF3",
            foreground=text,
            font=("Segoe UI", 9, "bold"),
            bordercolor=button_border,
            lightcolor="#E2EAF3",
            darkcolor=button_border,
            relief="flat",
        )
        style.map(
            "Treeview.Heading",
            background=[("active", "#D6E2EE"), ("pressed", "#C6D7E8")],
            bordercolor=[("active", "#7C93AA")],
            lightcolor=[("active", "#D6E2EE")],
            darkcolor=[("active", "#7C93AA")],
        )
        style.map("Treeview", background=[("selected", "#2F6F9F")], foreground=[("selected", "#FFFFFF")])
        style.configure(
            "Horizontal.TProgressbar",
            background=accent,
            troughcolor="#D8E1EC",
            bordercolor="#D8E1EC",
            lightcolor=accent,
            darkcolor=accent,
        )

        header = ttk.Frame(self.root, padding=(18, 16, 18, 10))
        header.pack(fill="x")
        self.header_label = ttk.Label(header, text=self.tr("header_title"), style="Header.TLabel")
        self.header_label.pack(side="left")
        controls = ttk.Frame(header)
        controls.pack(side="right")
        self.form_label = ttk.Label(controls, text=self.tr("form"))
        self.form_label.pack(side="left", padx=(0, 6))
        form_combo = ttk.Combobox(
            controls,
            textvariable=self.form_version,
            values=["USA", "Korea"],
            state="readonly",
            width=10,
        )
        form_combo.pack(side="left", padx=(0, 12))
        form_combo.bind("<<ComboboxSelected>>", lambda _event: self._on_form_version_changed())
        self.language_label = ttk.Label(controls, text=self.tr("language"))
        self.language_label.pack(side="left", padx=(0, 6))
        language_combo = ttk.Combobox(
            controls,
            textvariable=self.language_display,
            values=list(LANGUAGE_LABELS.values()),
            state="readonly",
            width=10,
        )
        language_combo.pack(side="left", padx=(0, 12))
        language_combo.bind("<<ComboboxSelected>>", lambda _event: self._on_language_changed())
        self.upload_folder_btn = ttk.Button(controls, text=self.tr("select_files"), command=self.select_files)
        self.upload_folder_btn.pack(side="left", padx=4)
        self.select_payment_proof_btn = ttk.Button(
            controls,
            text=self.tr("select_payment_proof"),
            command=self.select_support_files,
        )
        self.select_payment_proof_btn.pack(side="left", padx=4)
        self.generate_details_btn = ttk.Button(controls, text=self.tr("generate_details"), command=self.generate_selected_details)
        self.generate_details_btn.pack(side="left", padx=4)
        self.generate_all_btn = ttk.Button(controls, text=self.tr("generate_all"), command=self.generate_all_details)
        self.generate_all_btn.pack(side="left", padx=4)
        self.generate_excel_btn = ttk.Button(controls, text=self.tr("generate_excel"), command=self.generate_excel)
        self.generate_excel_btn.pack(side="left", padx=4)

        body = ttk.Frame(self.root, padding=(14, 0, 14, 10))
        body.pack(fill="both", expand=True)
        body.grid_columnconfigure(0, minsize=650, weight=0)
        body.grid_columnconfigure(1, weight=1)
        body.grid_rowconfigure(0, weight=1)

        manager = ttk.Frame(body, style="Panel.TFrame", padding=12)
        manager.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        manager.grid_columnconfigure(0, minsize=270, weight=1)
        manager.grid_columnconfigure(1, minsize=350, weight=1)
        manager.grid_rowconfigure(1, weight=1)
        self.manager_label = ttk.Label(manager, text=self.tr("inserted_receipts"), style="PanelHeader.TLabel")
        self.manager_label.grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 8)
        )

        left = ttk.Frame(manager, style="Panel.TFrame", padding=(0, 0, 10, 0))
        left.grid(row=1, column=0, sticky="nsew")
        self.tree = ttk.Treeview(
            left,
            columns=("status", "date", "amount"),
            show="tree headings",
            height=20,
            selectmode="extended",
        )
        self.tree.heading("#0", text=self.tr("file"))
        self.tree.heading("status", text=self.tr("status"))
        self.tree.heading("date", text=self.tr("date"))
        self.tree.heading("amount", text=self.tr("amount"))
        self.tree.column("#0", width=140, stretch=True)
        self.tree.column("status", width=78, stretch=False)
        self.tree.column("date", width=78, stretch=False)
        self.tree.column("amount", width=70, stretch=False)
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self.tree.bind("<Shift-Button-1>", self._on_tree_shift_click)
        self.tree.bind("<Delete>", self.remove_selected)
        self.tree.bind("<BackSpace>", self.remove_selected)
        self.tree.bind("<Control-a>", self.select_all_receipts)
        self.tree.bind("<Control-A>", self.select_all_receipts)
        left_buttons = ttk.Frame(left, style="Panel.TFrame")
        left_buttons.pack(fill="x", pady=(10, 0))
        self.remove_btn = ttk.Button(left_buttons, text=self.tr("remove"), command=self.remove_selected)
        self.remove_btn.pack(side="left")
        self.clear_btn = ttk.Button(left_buttons, text=self.tr("clear"), command=self.clear_all)
        self.clear_btn.pack(side="left", padx=(8, 0))

        middle = ttk.Frame(manager, style="Panel.TFrame", padding=(10, 0, 0, 0))
        middle.grid(row=1, column=1, sticky="nsew")
        self.details_label = ttk.Label(middle, text=self.tr("details"), style="PanelHeader.TLabel")
        self.details_label.grid(row=0, column=0, columnspan=2, sticky="w")
        rate_frame = ttk.Frame(middle, style="Panel.TFrame")
        rate_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(8, 10))
        self.usa_rate_label = ttk.Label(rate_frame, text="USD -> RMB", style="Panel.TLabel")
        self.usa_rate_label.pack(side="left")
        self.usa_rate_entry = ttk.Entry(rate_frame, textvariable=self.exchange_rate, width=9)
        self.usa_rate_entry.pack(side="left", padx=(6, 14))
        self.usd_krw_rate_label = ttk.Label(rate_frame, text="USD -> KRW", style="Panel.TLabel")
        self.usd_krw_rate_label.pack(side="left")
        self.usd_krw_rate_entry = ttk.Entry(rate_frame, textvariable=self.usd_to_krw_rate, width=9)
        self.usd_krw_rate_entry.pack(side="left", padx=(6, 14))
        self.krw_rate_label = ttk.Label(rate_frame, text="KRW -> RMB", style="Panel.TLabel")
        self.krw_rate_label.pack(side="left")
        self.krw_rate_entry = ttk.Entry(rate_frame, textvariable=self.krw_to_rmb_rate, width=9)
        self.krw_rate_entry.pack(side="left", padx=(6, 0))

        fields = [
            ("date", self.tr("date")),
            ("place", self.tr("place_vendor")),
            ("amount", self.tr("usd_amount")),
            ("currency", self.tr("currency")),
            ("krw_amount", self.tr("krw_amount")),
            ("rmb_amount", self.tr("rmb_amount")),
            ("purpose", self.tr("purpose")),
            ("details", self.tr("details")),
            ("project_number", self.tr("project_number")),
            ("category", self.tr("category")),
            ("payment_method", self.tr("payment_method")),
            ("receipt_label", self.tr("receipt_label")),
        ]
        for row, (key, label) in enumerate(fields, start=2):
            label_widget = ttk.Label(middle, text=label, style="Panel.TLabel")
            label_widget.grid(row=row, column=0, sticky="w", pady=4)
            if key == "category":
                var = tk.StringVar()
                widget = ttk.Combobox(middle, textvariable=var, state="readonly", width=28)
                self.category_combo = widget
            elif key == "currency":
                var = tk.StringVar(value="USD")
                widget = ttk.Combobox(middle, textvariable=var, values=["USD", "KRW", "RMB", "CNY"], width=28)
            else:
                var = tk.StringVar()
                widget = ttk.Entry(middle, textvariable=var, width=31)
            self.field_vars[key] = var
            self.field_labels[key] = label_widget
            self.field_widgets[key] = widget
            widget.grid(row=row, column=1, sticky="ew", pady=4, padx=(8, 0))
        middle.grid_columnconfigure(1, weight=1)

        right = ttk.Frame(body, style="Panel.TFrame", padding=12)
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_rowconfigure(1, weight=1)
        right.grid_columnconfigure(0, weight=1)
        preview_header = ttk.Frame(right, style="Panel.TFrame")
        preview_header.grid(row=0, column=0, sticky="ew")
        preview_header.grid_columnconfigure(0, weight=1)
        self.preview_label = ttk.Label(preview_header, text=self.tr("receipt_preview"), style="PanelHeader.TLabel")
        self.preview_label.grid(row=0, column=0, sticky="w")
        self.rotate_left_btn = ttk.Button(
            preview_header,
            text="\u27F2",
            command=lambda: self.rotate_selected(-90),
            state="disabled",
            style="Icon.TButton",
        )
        self.rotate_left_btn.grid(row=0, column=1, sticky="e", padx=(0, 6))
        ToolTip(self.rotate_left_btn, self.tr("rotate_left"))
        self.rotate_right_btn = ttk.Button(
            preview_header,
            text="\u27F3",
            command=lambda: self.rotate_selected(90),
            state="disabled",
            style="Icon.TButton",
        )
        self.rotate_right_btn.grid(row=0, column=2, sticky="e", padx=(0, 6))
        ToolTip(self.rotate_right_btn, self.tr("rotate_right"))
        self.revert_crop_btn = ttk.Button(
            preview_header,
            text=self.tr("revert"),
            command=self.revert_crop,
            state="disabled",
        )
        self.revert_crop_btn.grid(row=0, column=3, sticky="e")
        ToolTip(self.revert_crop_btn, self.tr("revert"))
        self.delete_screenshot_btn = ttk.Button(
            preview_header,
            text="\u2715",
            command=self.delete_selected_screenshot,
            state="disabled",
            style="Icon.TButton",
        )
        self.delete_screenshot_btn.grid(row=0, column=4, sticky="e", padx=(6, 0))
        ToolTip(self.delete_screenshot_btn, self.tr("delete_screenshot"))
        preview_frame = ttk.Frame(right, style="Panel.TFrame")
        preview_frame.grid(row=1, column=0, sticky="nsew", pady=(10, 0))
        preview_frame.grid_rowconfigure(0, weight=1)
        preview_frame.grid_columnconfigure(0, weight=1)
        self.preview_canvas = tk.Canvas(
            preview_frame,
            bg=canvas_bg,
            highlightthickness=1,
            highlightbackground=border,
            bd=1,
            relief="solid",
        )
        self.preview_canvas.grid(row=0, column=0, sticky="nsew")
        self.preview_canvas.bind("<Configure>", lambda _event: self.update_preview())
        self.preview_canvas.bind("<ButtonPress-1>", self._on_crop_press)
        self.preview_canvas.bind("<B1-Motion>", self._on_crop_drag)
        self.preview_canvas.bind("<ButtonRelease-1>", self._on_crop_release)

        footer = ttk.Frame(self.root, padding=(18, 0, 18, 12))
        footer.pack(fill="x")
        ttk.Label(footer, textvariable=self.status_text, style="Muted.TLabel").pack(side="left")
        ttk.Label(footer, textvariable=self.progress_text, style="Muted.TLabel").pack(side="left", padx=(16, 6))
        self.progress_bar = ttk.Progressbar(
            footer,
            variable=self.progress_value,
            maximum=100,
            length=180,
            mode="determinate",
        )
        self.progress_bar.pack(side="left")
        self.open_output_btn = ttk.Button(footer, text=self.tr("open_output_folder"), command=self.open_output_folder)
        self.open_output_btn.pack(side="right")

    def _attach_amount_traces(self) -> None:
        for key in ("amount", "currency", "krw_amount", "rmb_amount"):
            var = self.field_vars.get(key)
            if var is not None:
                var.trace_add("write", lambda *_args, k=key: self._sync_amount_fields(k))
        for key in (
            "date",
            "place",
            "purpose",
            "details",
            "project_number",
            "category",
            "payment_method",
            "receipt_label",
        ):
            var = self.field_vars.get(key)
            if var is not None:
                var.trace_add("write", lambda *_args, k=key: self._on_field_changed(k))
        self.exchange_rate.trace_add("write", lambda *_args: self._sync_amount_fields("rate"))
        self.usd_to_krw_rate.trace_add("write", lambda *_args: self._sync_amount_fields("rate"))
        self.krw_to_rmb_rate.trace_add("write", lambda *_args: self._sync_amount_fields("rate"))

    def category_display_label(self, category: str) -> str:
        labels = CATEGORY_LABELS_BY_LANGUAGE.get(self.language.get(), CATEGORY_LABELS_BY_LANGUAGE["en"])
        return labels.get(category, category)

    def category_key_from_display(self, value: str) -> str:
        return self.category_label_to_key.get(value, category_value_to_key(value, self.form_version.get()))

    def _on_language_changed(self) -> None:
        chosen_label = self.language_display.get()
        for code, label in LANGUAGE_LABELS.items():
            if label == chosen_label:
                self.language.set(code)
                break
        else:
            self.language.set("en")
            self.language_display.set(LANGUAGE_LABELS["en"])
        self.settings["language"] = self.language.get()
        save_user_settings(self.settings)
        self.refresh_language_texts()

    def refresh_language_texts(self) -> None:
        self.root.title(self.tr("app_title"))
        for widget, key in (
            (self.header_label, "header_title"),
            (self.form_label, "form"),
            (self.language_label, "language"),
            (self.upload_folder_btn, "select_files"),
            (self.generate_details_btn, "generate_details"),
            (self.generate_all_btn, "generate_all"),
            (self.generate_excel_btn, "generate_excel"),
            (self.manager_label, "inserted_receipts"),
            (self.details_label, "details"),
            (self.preview_label, "receipt_preview"),
            (self.remove_btn, "remove"),
            (self.clear_btn, "clear"),
            (self.revert_crop_btn, "revert"),
            (self.open_output_btn, "open_output_folder"),
        ):
            if widget is not None:
                try:
                    widget.configure(text=self.tr(key))
                except Exception:
                    pass
        self.tree.heading("#0", text=self.tr("file"))
        self.tree.heading("status", text=self.tr("status"))
        self.tree.heading("date", text=self.tr("date"))
        self.tree.heading("amount", text=self.tr("amount"))
        for key, text_key in {
            "date": "date",
            "place": "place_vendor",
            "currency": "currency",
            "krw_amount": "krw_amount",
            "rmb_amount": "rmb_amount",
            "purpose": "purpose",
            "details": "details",
            "project_number": "project_number",
            "category": "category",
            "payment_method": "payment_method",
            "receipt_label": "receipt_label",
        }.items():
            label = self.field_labels.get(key)
            if label is not None:
                label.configure(text=self.tr(text_key))
        self._update_field_visibility()
        self._on_form_version_changed(refresh_language_only=True)
        if self.status_text.get() in {ui_text(code, "ready") for code in UI_TEXT}:
            self.status_text.set(self.tr("ready"))
        self.update_preview()

    def _on_field_changed(self, key: str) -> None:
        if self._loading_fields or self._syncing_amounts:
            return
        self._save_field_to_selected(key, refresh=True)

    def _sync_amount_fields(self, changed_key: str) -> None:
        if self._loading_fields or self._syncing_amounts:
            return
        self._syncing_amounts = True
        should_save = False
        try:
            version = self.form_version.get()
            amount_var = self.field_vars.get("amount")
            currency_var = self.field_vars.get("currency")
            krw_var = self.field_vars.get("krw_amount")
            rmb_var = self.field_vars.get("rmb_amount")
            if amount_var is None or currency_var is None or krw_var is None or rmb_var is None:
                should_save = False
            if changed_key in {"amount", "currency", "krw_amount", "rmb_amount"}:
                self._last_amount_source = changed_key

            if amount_var is None or currency_var is None or krw_var is None or rmb_var is None:
                return

            if version == "USA":
                rate = safe_float(self.exchange_rate.get())
                if rate:
                    if currency_var.get() != "USD":
                        currency_var.set("USD")
                    source = self._last_amount_source if changed_key == "rate" else changed_key
                    if source == "rmb_amount":
                        rmb = safe_float(rmb_var.get())
                        if rmb is not None:
                            amount_var.set(format_amount(rmb / rate))
                    else:
                        usd = safe_float(amount_var.get())
                        if usd is not None:
                            rmb_var.set(format_amount(usd * rate))
                    should_save = True
            else:
                krw_rate = safe_float(self.krw_to_rmb_rate.get())
                usd_rmb_rate = safe_float(self.exchange_rate.get())
                usd_krw_rate = safe_float(self.usd_to_krw_rate.get())
                if krw_rate:
                    currency = normalize_currency(currency_var.get(), "KRW")
                    if currency_var.get() != currency:
                        currency_var.set(currency)
                    source = self._last_amount_source if changed_key == "rate" else changed_key
                    amount = safe_float(amount_var.get())
                    krw = safe_float(krw_var.get())
                    rmb = safe_float(rmb_var.get())
                    if source in {"amount", "currency", "rate"} and amount is not None:
                        if currency == "USD":
                            if usd_krw_rate:
                                krw = amount * usd_krw_rate
                                rmb = krw * krw_rate
                            elif usd_rmb_rate:
                                rmb = amount * usd_rmb_rate
                                krw = rmb / krw_rate
                        elif currency in {"RMB", "CNY"}:
                            rmb = amount
                            krw = amount / krw_rate
                        else:
                            krw = amount
                            rmb = amount * krw_rate
                    elif source == "rmb_amount" and rmb is not None:
                        krw = rmb / krw_rate
                        if currency in {"RMB", "CNY"}:
                            amount_var.set(format_amount(rmb))
                        elif currency == "KRW":
                            amount_var.set(format_amount(krw))
                        elif currency == "USD" and usd_krw_rate and not amount_var.get().strip():
                            amount_var.set(format_amount(krw / usd_krw_rate))
                        elif currency == "USD" and usd_rmb_rate and not amount_var.get().strip():
                            amount_var.set(format_amount(rmb / usd_rmb_rate))
                    elif source == "krw_amount" and krw is not None:
                        rmb = krw * krw_rate
                        if currency == "KRW":
                            amount_var.set(format_amount(krw))
                        elif currency in {"RMB", "CNY"} and not amount_var.get().strip():
                            amount_var.set(format_amount(rmb))
                        elif currency == "USD" and usd_krw_rate and not amount_var.get().strip():
                            amount_var.set(format_amount(krw / usd_krw_rate))
                        elif currency == "USD" and usd_rmb_rate and not amount_var.get().strip():
                            amount_var.set(format_amount(rmb / usd_rmb_rate))

                    if krw is None and rmb is not None:
                        krw = rmb / krw_rate
                    if rmb is None and krw is not None:
                        rmb = krw * krw_rate
                    if krw is not None:
                        krw_var.set(format_amount(krw))
                    if rmb is not None:
                        rmb_var.set(format_amount(rmb))
                    should_save = True
        finally:
            self._syncing_amounts = False
        if should_save:
            self._save_amount_fields_to_selected(changed_key, refresh=True)

    def update_item_amount_fields(self, item: ReceiptItem, source: str = "amount") -> None:
        version = self.form_version.get()
        if version == "USA":
            item.currency = "USD"
            rate = safe_float(self.exchange_rate.get()) or safe_float(DEFAULT_USD_TO_RMB_RATE)
            if not rate:
                return
            amount = safe_float(item.amount)
            rmb = safe_float(item.rmb_amount)
            if source == "rmb_amount" and rmb is not None:
                item.amount = format_amount(rmb / rate)
            elif amount is not None:
                item.rmb_amount = format_amount(amount * rate)
            return

        krw_rate = safe_float(self.krw_to_rmb_rate.get()) or safe_float(DEFAULT_KRW_TO_RMB_RATE)
        usd_rmb_rate = safe_float(self.exchange_rate.get()) or safe_float(DEFAULT_USD_TO_RMB_RATE)
        usd_krw_rate = safe_float(self.usd_to_krw_rate.get()) or safe_float(DEFAULT_USD_TO_KRW_RATE)
        if not krw_rate:
            return
        item.currency = normalize_currency(item.currency, "KRW")
        amount = safe_float(item.amount)
        krw = safe_float(item.krw_amount)
        rmb = safe_float(item.rmb_amount)
        item_source = "amount" if source in {"currency", "rate"} and amount is not None else source
        if item_source == "amount" and amount is not None:
            if item.currency == "USD":
                krw = amount * usd_krw_rate if usd_krw_rate else krw
                rmb = krw * krw_rate if krw is not None else (amount * usd_rmb_rate if usd_rmb_rate else rmb)
            elif item.currency in {"RMB", "CNY"}:
                rmb = amount
                krw = amount / krw_rate
            else:
                krw = amount
                rmb = amount * krw_rate
        elif item_source == "rmb_amount" and rmb is not None:
            krw = rmb / krw_rate
            if item.currency in {"RMB", "CNY"}:
                item.amount = format_amount(rmb)
            elif item.currency == "KRW" and amount is None:
                item.amount = format_amount(krw)
            elif item.currency == "USD" and amount is None and usd_krw_rate:
                item.amount = format_amount(krw / usd_krw_rate)
            elif item.currency == "USD" and amount is None and usd_rmb_rate:
                item.amount = format_amount(rmb / usd_rmb_rate)
        elif item_source == "krw_amount" and krw is not None:
            rmb = krw * krw_rate
            if item.currency == "KRW":
                item.amount = format_amount(krw)
            elif item.currency in {"RMB", "CNY"} and amount is None:
                item.amount = format_amount(rmb)
            elif item.currency == "USD" and amount is None and usd_krw_rate:
                item.amount = format_amount(krw / usd_krw_rate)
            elif item.currency == "USD" and amount is None and usd_rmb_rate:
                item.amount = format_amount(rmb / usd_rmb_rate)

        if krw is None and rmb is not None:
            krw = rmb / krw_rate
        if rmb is None and krw is not None:
            rmb = krw * krw_rate
        if krw is not None:
            item.krw_amount = format_amount(krw)
        if rmb is not None:
            item.rmb_amount = format_amount(rmb)

    def _save_field_to_selected(self, key: str, refresh: bool = False) -> None:
        var = self.field_vars.get(key)
        if var is None:
            return
        value = var.get()
        version = self.form_version.get()
        for index in self.selected_indices():
            item = self.items[index]
            if key == "category":
                setattr(item, key, self.category_key_from_display(value))
            else:
                setattr(item, key, value)
        if refresh:
            self.refresh_tree()

    def _save_amount_fields_to_selected(self, source: str, refresh: bool = False) -> None:
        version = self.form_version.get()
        amount_value = self.field_vars["amount"].get()
        currency_value = normalize_currency(self.field_vars["currency"].get(), "USD" if version == "USA" else "KRW")
        krw_value = self.field_vars["krw_amount"].get()
        rmb_value = self.field_vars["rmb_amount"].get()
        indices = range(len(self.items)) if source == "rate" else self.selected_indices()
        for index in indices:
            item = self.items[index]
            if source == "rate":
                self.update_item_amount_fields(item, "amount")
                continue
            if source == "amount":
                item.amount = amount_value
                if version == "USA":
                    item.currency = "USD"
                self.update_item_amount_fields(item, "amount")
            elif source == "currency":
                item.currency = currency_value
                self.update_item_amount_fields(item, "currency")
            elif source == "krw_amount":
                item.krw_amount = krw_value
                self.update_item_amount_fields(item, "krw_amount")
            elif source == "rmb_amount":
                item.rmb_amount = rmb_value
                self.update_item_amount_fields(item, "rmb_amount")
        if refresh:
            self.refresh_tree()

    def _set_field_visible(self, key: str, visible: bool) -> None:
        label = self.field_labels.get(key)
        widget = self.field_widgets.get(key)
        if visible:
            if label is not None:
                label.grid()
            if widget is not None:
                widget.grid()
        else:
            if label is not None:
                label.grid_remove()
            if widget is not None:
                widget.grid_remove()

    def _update_field_visibility(self) -> None:
        version = self.form_version.get()
        if version == "USA":
            self.field_labels["amount"].configure(text=self.tr("usd_amount"))
            self.field_labels["rmb_amount"].configure(text=self.tr("rmb_amount"))
            self.field_vars["currency"].set("USD")
            self._set_field_visible("currency", False)
            self._set_field_visible("krw_amount", False)
            self._set_field_visible("amount", True)
            self._set_field_visible("rmb_amount", True)
            if self.usa_rate_label is not None:
                self.usa_rate_label.pack(side="left")
            if self.usa_rate_entry is not None:
                self.usa_rate_entry.pack(side="left", padx=(6, 14))
            if self.usd_krw_rate_label is not None:
                self.usd_krw_rate_label.pack_forget()
            if self.usd_krw_rate_entry is not None:
                self.usd_krw_rate_entry.pack_forget()
            if self.krw_rate_label is not None:
                self.krw_rate_label.pack_forget()
            if self.krw_rate_entry is not None:
                self.krw_rate_entry.pack_forget()
            if self.select_payment_proof_btn is not None:
                self.select_payment_proof_btn.configure(text=self.tr("select_payment_proof"), command=self.select_support_files, state="normal")
        else:
            self.field_labels["amount"].configure(text=self.tr("original_amount"))
            self.field_labels["rmb_amount"].configure(text=self.tr("rmb_amount"))
            self._set_field_visible("currency", True)
            self._set_field_visible("krw_amount", True)
            self._set_field_visible("amount", True)
            self._set_field_visible("rmb_amount", True)
            if self.usa_rate_label is not None:
                self.usa_rate_label.pack_forget()
            if self.usa_rate_entry is not None:
                self.usa_rate_entry.pack_forget()
            if self.usd_krw_rate_label is not None:
                self.usd_krw_rate_label.pack(side="left")
            if self.usd_krw_rate_entry is not None:
                self.usd_krw_rate_entry.pack(side="left", padx=(6, 14))
            if self.krw_rate_label is not None:
                self.krw_rate_label.pack(side="left")
            if self.krw_rate_entry is not None:
                self.krw_rate_entry.pack(side="left", padx=(6, 0))
            if self.select_payment_proof_btn is not None:
                self.select_payment_proof_btn.configure(text=self.tr("select_exchange_rate"), command=self.select_support_files, state="normal")

    def suggested_toolbar_action(self) -> str:
        if self._busy:
            return ""
        if self.items and self._details_ready_for_export:
            return "generate_excel"
        if not self.items:
            return "select_files"
        if self.form_version.get() == "USA" and not self.bank_items:
            return "select_payment_proof"
        if self.form_version.get() == "Korea" and not self.exchange_rate_items:
            return "select_payment_proof"
        return "generate_all"

    def update_toolbar_recommendation(self) -> None:
        suggested = self.suggested_toolbar_action()
        buttons = {
            "select_files": self.upload_folder_btn,
            "select_payment_proof": self.select_payment_proof_btn,
            "generate_all": self.generate_all_btn,
            "generate_excel": self.generate_excel_btn,
        }
        for action, button in buttons.items():
            if button is None:
                continue
            try:
                button.configure(style="Recommended.TButton" if action == suggested else "TButton")
            except Exception:
                pass

    def _on_form_version_changed(self, refresh_language_only: bool = False) -> None:
        version = normalized_form_version(self.form_version.get())
        if version != self.form_version.get():
            self.form_version.set(version)
        form_changed = version != self._last_form_version and not refresh_language_only
        if form_changed:
            self._details_ready_for_export = False
            self.settings["last_form_version"] = version
            save_user_settings(self.settings)
        cat_var = self.field_vars.get("category")
        current_key = self.category_key_from_display(cat_var.get()) if cat_var is not None else ""
        order = KOREA_REPORT_CATEGORY_ORDER if version == "Korea" else USA_REPORT_CATEGORY_ORDER
        self.category_values = [self.category_display_label(key) for key in order]
        self.category_label_to_key = {
            self.category_display_label(key): key
            for key in order
        }
        if self.category_combo is not None:
            self.category_combo.configure(values=self.category_values)
        cat = self.field_vars.get("category")
        if cat is not None:
            if current_key in order:
                cat.set(self.category_display_label(current_key))
            elif not cat.get():
                cat.set(self.category_values[0] if self.category_values else "transportation")
        for item in self.items:
            if version == "USA":
                item.currency = "USD"
            else:
                item.currency = normalize_currency(item.currency, "KRW")
            self.update_item_amount_fields(item, "amount")
            if form_changed and item.status == "AI filled":
                item.status = "Needs regen"
        self._update_field_visibility()
        if self.selected_index is not None:
            self.load_selected_into_fields()
        self._sync_amount_fields("rate")
        self.refresh_tree()
        self.refresh_bank_tree()
        self._last_form_version = version
        self.update_toolbar_recommendation()

    def select_files(self) -> None:
        ensure_runtime_folders()
        if filedialog is None:
            return
        selected = filedialog.askopenfilenames(
            title="Select receipt files",
            filetypes=[
                ("Receipt files", "*.png *.jpg *.jpeg *.webp *.bmp *.gif *.pdf"),
                ("Image files", "*.png *.jpg *.jpeg *.webp *.bmp *.gif"),
                ("PDF files", "*.pdf"),
                ("All files", "*.*"),
            ],
        )
        paths = [Path(path) for path in selected]
        if not paths:
            return
        self.save_current_fields()
        self._details_ready_for_export = False
        existing = {
            selected_file_key(Path(item.path), item.source_path, item.source_page)
            for item in [*self.items, *self.bank_items]
            if item.path
        }
        added = 0
        failed: List[str] = []
        first_new_index = len(self.items)
        for path in paths:
            suffix = path.suffix.lower()
            if suffix not in SUPPORTED_UPLOAD_SUFFIXES:
                continue
            if suffix == ".pdf":
                try:
                    merged_path = render_pdf_to_merged_image(path)
                except Exception as exc:
                    log_exception(f"Could not import PDF: {path}")
                    failed.append(f"{path.name}: {exc}")
                    continue
                key = selected_file_key(merged_path, str(path), "merged")
                if key in existing:
                    continue
                filename = f"{path.name} (all pages)"
                self.items.append(
                    ReceiptItem(
                        item_id=uuid.uuid4().hex,
                        path=str(merged_path),
                        filename=filename,
                        source_path=str(path),
                        source_page="merged",
                        currency="USD" if self.form_version.get() == "USA" else "KRW",
                        category="transportation",
                        receipt_images=[
                            {
                                "id": uuid.uuid4().hex,
                                "path": str(merged_path),
                                "filename": filename,
                                "source_path": str(path),
                                "source_page": "merged",
                                "crop_box": None,
                                "rotation_degrees": 0,
                            }
                        ],
                    )
                )
                existing.add(key)
                added += 1
                continue
            key = selected_file_key(path, str(path), "")
            if key in existing:
                continue
            self.items.append(
                ReceiptItem(
                    item_id=uuid.uuid4().hex,
                    path=str(path),
                    filename=path.name,
                    source_path=str(path),
                    currency="USD" if self.form_version.get() == "USA" else "KRW",
                    category="transportation",
                    receipt_images=[
                        {
                            "id": uuid.uuid4().hex,
                            "path": str(path),
                            "filename": path.name,
                            "source_path": str(path),
                            "source_page": "",
                            "crop_box": None,
                            "rotation_degrees": 0,
                        }
                    ],
                )
            )
            existing.add(key)
            added += 1
        self.refresh_tree()
        if added and self.items and self.selected_index is None:
            self.select_index(first_new_index)
        self.save_session()
        self.status_text.set(f"Added {added} receipt file(s).")
        self.update_toolbar_recommendation()
        if failed and messagebox:
            messagebox.showwarning(
                "Select Files",
                "Some files could not be imported.\n\n"
                + "\n".join(failed[:5])
                + f"\n\nLogged to:\n{LOG_DIR / 'app.log'}",
            )

    def upload_folder(self) -> None:
        self.select_files()

    def select_support_files(self) -> None:
        if self.form_version.get() == "USA":
            self.select_payment_proofs()
        else:
            self.select_exchange_rate_images()

    def select_payment_proofs(self) -> None:
        if self.form_version.get() != "USA":
            self.show_info("Payment proof images are only used for the USA form.")
            return
        ensure_runtime_folders()
        if filedialog is None:
            return
        selected = filedialog.askopenfilenames(
            title="Select payment proof files",
            filetypes=[
                ("Payment proof files", "*.png *.jpg *.jpeg *.webp *.bmp *.gif *.pdf"),
                ("Image files", "*.png *.jpg *.jpeg *.webp *.bmp *.gif"),
                ("PDF files", "*.pdf"),
                ("All files", "*.*"),
            ],
        )
        paths = [Path(path) for path in selected]
        if not paths:
            return
        self._details_ready_for_export = False
        existing = {
            selected_file_key(Path(item.path), item.source_path, item.source_page)
            for item in self.bank_items
            if item.path
        }
        added = 0
        failed: List[str] = []
        for path in paths:
            suffix = path.suffix.lower()
            if suffix not in SUPPORTED_UPLOAD_SUFFIXES:
                continue
            if suffix == ".pdf":
                try:
                    merged_path = render_pdf_to_merged_image(path)
                except Exception as exc:
                    log_exception(f"Could not import payment proof PDF: {path}")
                    failed.append(f"{path.name}: {exc}")
                    continue
                key = selected_file_key(merged_path, str(path), "merged")
                if key in existing:
                    continue
                self.bank_items.append(
                    BankStatementItem(
                        item_id=uuid.uuid4().hex,
                        path=str(merged_path),
                        filename=f"{path.name} (all pages)",
                        source_path=str(path),
                        source_page="merged",
                    )
                )
                existing.add(key)
                added += 1
                continue
            key = selected_file_key(path, str(path), "")
            if key in existing:
                continue
            self.bank_items.append(
                BankStatementItem(
                    item_id=uuid.uuid4().hex,
                    path=str(path),
                    filename=path.name,
                    source_path=str(path),
                )
            )
            existing.add(key)
            added += 1
        self.save_session()
        self.status_text.set(f"Added {added} payment proof file(s).")
        self.update_toolbar_recommendation()
        if self.selected_item() is not None:
            self.update_preview()
        if failed and messagebox:
            messagebox.showwarning(
                "Select Payment Proof",
                "Some files could not be imported.\n\n"
                + "\n".join(failed[:5])
                + f"\n\nLogged to:\n{LOG_DIR / 'app.log'}",
            )

    def select_exchange_rate_images(self) -> None:
        if self.form_version.get() != "Korea":
            self.show_info(self.tr("exchange_only_korea"))
            return
        ensure_runtime_folders()
        if filedialog is None:
            return
        selected = filedialog.askopenfilenames(
            title=self.tr("select_exchange_rate"),
            filetypes=[
                ("Exchange rate image files", "*.png *.jpg *.jpeg *.webp *.bmp *.gif *.pdf"),
                ("Image files", "*.png *.jpg *.jpeg *.webp *.bmp *.gif"),
                ("PDF files", "*.pdf"),
                ("All files", "*.*"),
            ],
        )
        paths = [Path(path) for path in selected]
        if not paths:
            return
        self._details_ready_for_export = False
        self.exchange_rate_items.clear()
        added = 0
        failed: List[str] = []
        for path in paths:
            suffix = path.suffix.lower()
            if suffix not in SUPPORTED_UPLOAD_SUFFIXES:
                continue
            if suffix == ".pdf":
                try:
                    image_path = render_pdf_to_merged_image(path)
                except Exception as exc:
                    log_exception(f"Could not import 汇率 PDF: {path}")
                    failed.append(f"{path.name}: {exc}")
                    continue
                filename = f"{path.name} (all pages)"
                source_page = "merged"
            else:
                image_path = path
                filename = path.name
                source_page = ""
            self.exchange_rate_items.append(
                BankStatementItem(
                    item_id=uuid.uuid4().hex,
                    path=str(image_path),
                    filename=filename,
                    source_path=str(path),
                    source_page=source_page,
                    status="Exchange rate",
                )
            )
            added += 1
        self.save_session()
        self.status_text.set(self.tr("selected_exchange_files", count=added))
        self.update_toolbar_recommendation()
        if added:
            self.read_exchange_rate_from_images()
        if failed and messagebox:
            messagebox.showwarning(
                self.tr("select_exchange_rate"),
                "Some files could not be imported.\n\n"
                + "\n".join(failed[:5])
                + f"\n\nLogged to:\n{LOG_DIR / 'app.log'}",
            )

    def read_exchange_rate_from_images(self) -> None:
        if not self.exchange_rate_items:
            return
        if not get_openai_api_key():
            self.status_text.set(self.tr("selected_exchange_no_key"))
            return
        image_paths = [Path(item.path) for item in self.exchange_rate_items]
        usd_to_rmb_rate = safe_float(self.exchange_rate.get()) or safe_float(DEFAULT_USD_TO_RMB_RATE) or 6.8175
        self.set_busy(True)
        self.set_progress(0, 1, self.tr("reading_exchange_short"))
        self.status_text.set(self.tr("reading_exchange"))

        def worker() -> None:
            try:
                rates = call_openai_exchange_rate_extraction(
                    image_paths,
                    usd_to_rmb_rate,
                    progress=lambda message: self.set_status(message),
                )
            except Exception as exc:
                log_exception("Could not read exchange rate images")
                message = str(exc)

                def fail() -> None:
                    self.set_busy(False)
                    self.set_progress(0, 1, "")
                    self.status_text.set(self.tr("selected_exchange_failed", message=message))

                self.root.after(0, fail)
                return

            def done() -> None:
                updates: List[str] = []
                if "usd_to_krw_rate" in rates:
                    self.usd_to_krw_rate.set(format_exchange_rate(rates["usd_to_krw_rate"]))
                    self.settings["usd_to_krw_rate"] = self.usd_to_krw_rate.get()
                    updates.append(f"USD -> KRW {self.usd_to_krw_rate.get()}")
                if "krw_to_rmb_rate" in rates:
                    self.krw_to_rmb_rate.set(format_exchange_rate(rates["krw_to_rmb_rate"]))
                    self.settings["krw_to_rmb_rate"] = self.krw_to_rmb_rate.get()
                    updates.append(f"KRW -> RMB {self.krw_to_rmb_rate.get()}")
                save_user_settings(self.settings)
                self.save_session()
                self.set_busy(False)
                self.set_progress(1, 1, "1/1")
                self.status_text.set(self.tr("updated_exchange_rate", updates=", ".join(updates)))

            self.root.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def refresh_tree(self) -> None:
        current_selection = set(self.tree.selection())
        self.tree.delete(*self.tree.get_children())
        for index, item in enumerate(self.items):
            values = (item.status, item.date, item.amount or item.krw_amount or item.rmb_amount)
            self.tree.insert("", "end", iid=str(index), text=item.filename, values=values)
        valid_selection = [iid for iid in current_selection if iid.isdigit() and int(iid) < len(self.items)]
        if valid_selection:
            self.tree.selection_set(*valid_selection)
        elif self.selected_index is not None and self.selected_index < len(self.items):
            self.tree.selection_set(str(self.selected_index))
        if self.selected_index is not None and self.selected_index < len(self.items):
            self.tree.focus(str(self.selected_index))

    def receipt_label_by_id(self, item_id: str) -> str:
        for item in self.items:
            if item.item_id == item_id:
                return item.receipt_label.strip() or item.filename
        return ""

    def refresh_bank_tree(self) -> None:
        if self.bank_tree is None:
            return
        current_selection = set(self.bank_tree.selection())
        self.bank_tree.delete(*self.bank_tree.get_children())
        for index, item in enumerate(self.bank_items):
            values = (
                item.status,
                item.date,
                item.amount,
                self.receipt_label_by_id(item.matched_receipt_id),
            )
            self.bank_tree.insert("", "end", iid=str(index), text=item.filename, values=values)
        valid_selection = [iid for iid in current_selection if iid.isdigit() and int(iid) < len(self.bank_items)]
        if valid_selection:
            self.bank_tree.selection_set(*valid_selection)
        elif self.selected_bank_index is not None and self.selected_bank_index < len(self.bank_items):
            self.bank_tree.selection_set(str(self.selected_bank_index))

    def _on_bank_tree_select(self, _event: Any = None) -> None:
        if self.bank_tree is None:
            return
        selection = self.bank_tree.selection()
        if not selection:
            self.selected_bank_index = None
            return
        try:
            self.selected_bank_index = int(selection[0])
        except Exception:
            self.selected_bank_index = None
        if self.selected_bank_index is not None:
            self.selected_attachment_kind = "proof"
            self.selected_attachment_index = 0
            self.update_preview()

    def selected_bank_item(self) -> Optional[BankStatementItem]:
        if self.selected_bank_index is None:
            return None
        if self.selected_bank_index < 0 or self.selected_bank_index >= len(self.bank_items):
            return None
        return self.bank_items[self.selected_bank_index]

    def _on_receipt_drag_start(self, event: Any) -> None:
        return

    def _is_widget_inside_bank_section(self, widget: Any) -> bool:
        return False

    def _on_receipt_drag_release(self, event: Any) -> None:
        return

    def move_selected_receipts_to_bank(self) -> None:
        self.show_info("Use Select Payment Proof to add USA payment proof images.")

    def move_receipts_to_bank(self, indices: List[int]) -> None:
        self.show_info("Use Select Payment Proof to add USA payment proof images.")

    def remove_selected_bank(self) -> None:
        item = self.selected_bank_item()
        if item is None:
            return
        del self.bank_items[self.selected_bank_index]
        self.selected_bank_index = None
        self.refresh_bank_tree()
        self.selected_attachment_kind = "receipt"
        self.selected_attachment_index = 0
        self.update_preview()
        self.save_session()
        self.status_text.set("Removed selected payment proof image.")
        self._details_ready_for_export = False
        self.update_toolbar_recommendation()

    def link_selected_bank_to_receipt(self) -> None:
        bank_item = self.selected_bank_item()
        receipt = self.selected_item()
        if bank_item is None or receipt is None:
            self.show_info("Select a payment proof image and the matching receipt row first.")
            return
        bank_item.matched_receipt_id = receipt.item_id
        bank_item.status = "Matched manually"
        self.refresh_bank_tree()
        self.save_session()
        self.status_text.set("Linked payment proof to selected receipt.")
        self._details_ready_for_export = False
        self.update_toolbar_recommendation()

    def _on_tree_select(self, _event: Any = None) -> None:
        selection = self.tree.selection()
        if not selection:
            return
        self.selected_attachment_kind = "receipt"
        self.selected_attachment_index = 0
        focus = self.tree.focus()
        chosen = focus if focus in selection else selection[-1]
        try:
            index = int(chosen)
        except Exception:
            return
        self.select_index(index)
        if len(selection) == 1:
            self._selection_anchor_index = index

    def _on_tree_shift_click(self, event: Any) -> str:
        iid = self.tree.identify_row(event.y)
        if not iid:
            return ""
        try:
            index = int(iid)
        except Exception:
            return "break"
        if index < 0 or index >= len(self.items):
            return "break"
        anchor = self._selection_anchor_index
        if anchor is None or anchor < 0 or anchor >= len(self.items):
            anchor = self.selected_index if self.selected_index is not None else index
        if anchor is None or anchor < 0 or anchor >= len(self.items):
            anchor = index
        start, end = sorted((anchor, index))
        selection = [str(row) for row in range(start, end + 1)]
        self.save_current_fields()
        self.tree.selection_set(selection)
        self.tree.focus(str(index))
        self.selected_index = index
        self.selected_attachment_kind = "receipt"
        self.selected_attachment_index = 0
        self.load_selected_into_fields()
        self.update_preview()
        return "break"

    def select_index(self, index: int) -> None:
        if self.selected_index == index:
            return
        self.save_current_fields()
        self.selected_index = index
        self._selection_anchor_index = index
        self.tree.focus(str(index))
        if str(index) not in self.tree.selection():
            self.tree.selection_set(str(index))
        self.load_selected_into_fields()
        self.update_preview()

    def selected_indices(self) -> List[int]:
        indices: List[int] = []
        for iid in self.tree.selection():
            try:
                index = int(iid)
            except Exception:
                continue
            if 0 <= index < len(self.items):
                indices.append(index)
        if not indices and self.selected_index is not None and 0 <= self.selected_index < len(self.items):
            indices.append(self.selected_index)
        return sorted(set(indices))

    def select_all_receipts(self, _event: Any = None) -> str:
        if not self.items:
            return "break"
        self.tree.selection_set([str(index) for index in range(len(self.items))])
        self.tree.focus(str(0))
        self.selected_index = 0
        self._selection_anchor_index = 0
        self.load_selected_into_fields()
        self.update_preview()
        self.status_text.set(f"Selected all {len(self.items)} receipt row(s).")
        return "break"

    def selected_items(self) -> List[ReceiptItem]:
        return [self.items[index] for index in self.selected_indices()]

    def sort_items_for_current_report(self) -> None:
        if not self.items:
            return
        selected_ids = {
            self.items[index].item_id
            for index in self.selected_indices()
            if 0 <= index < len(self.items)
        }
        focused_id = self.selected_item().item_id if self.selected_item() is not None else ""
        self.items = sort_receipts_for_report(self.items, self.form_version.get())
        id_to_index = {item.item_id: index for index, item in enumerate(self.items)}
        if focused_id in id_to_index:
            self.selected_index = id_to_index[focused_id]
        elif selected_ids:
            self.selected_index = id_to_index.get(next(iter(selected_ids)), 0)
        elif self.items:
            self.selected_index = 0
        else:
            self.selected_index = None
        selection = [str(id_to_index[item_id]) for item_id in selected_ids if item_id in id_to_index]
        self.refresh_tree()
        if selection:
            self.tree.selection_set(selection)
        elif self.selected_index is not None:
            self.tree.selection_set(str(self.selected_index))

    def selected_item(self) -> Optional[ReceiptItem]:
        if self.selected_index is None:
            return None
        if self.selected_index < 0 or self.selected_index >= len(self.items):
            return None
        return self.items[self.selected_index]

    def save_current_fields(self) -> None:
        item = self.selected_item()
        if item is None:
            return
        for key, var in self.field_vars.items():
            value = var.get()
            if key == "category":
                value = self.category_key_from_display(value)
            elif key == "currency":
                value = normalize_currency(value, "USD" if self.form_version.get() == "USA" else "KRW")
            setattr(item, key, value)
        self.update_item_amount_fields(item, self._last_amount_source)

    def load_selected_into_fields(self) -> None:
        item = self.selected_item()
        if item is None:
            self._loading_fields = True
            try:
                for var in self.field_vars.values():
                    var.set("")
            finally:
                self._loading_fields = False
            return
        self._loading_fields = True
        try:
            for key, var in self.field_vars.items():
                value = getattr(item, key)
                if key == "category":
                    value = self.category_display_label(category_value_to_key(value, self.form_version.get()))
                var.set(value)
        finally:
            self._loading_fields = False
        self._sync_amount_fields("rate")

    def clear_preview(self, message: str) -> None:
        if self.preview_canvas is None:
            return
        self.preview_canvas.delete("all")
        width = max(1, self.preview_canvas.winfo_width())
        height = max(1, self.preview_canvas.winfo_height())
        self.preview_canvas.create_text(
            width / 2,
            height / 2,
            text=message,
            fill="#666666",
            width=max(160, width - 30),
            justify="center",
        )
        self.photo = None
        self.preview_photos = []
        self.preview_tiles = []
        self.preview_action_regions = {}
        self.proof_drop_region = None
        self._pressed_preview_tile = None
        self._dragging_preview_tile = None
        self._drag_start = (0, 0)
        self._drag_outline_id = None
        self._preview_image_bounds = (0, 0, 0, 0)
        self._preview_original_size = (0, 0)
        self._crop_handle_centers = {}
        if self.revert_crop_btn is not None:
            self.revert_crop_btn.configure(state="disabled")
        if self.rotate_left_btn is not None:
            self.rotate_left_btn.configure(state="disabled")
        if self.rotate_right_btn is not None:
            self.rotate_right_btn.configure(state="disabled")
        if self.delete_screenshot_btn is not None:
            self.delete_screenshot_btn.configure(state="disabled")
        if self.swap_proof_btn is not None:
            self.swap_proof_btn.configure(state="disabled")
        if self.unlink_proof_btn is not None:
            self.unlink_proof_btn.configure(state="disabled")

    def payment_proofs_for_receipt(self, receipt: ReceiptItem) -> List[BankStatementItem]:
        return [item for item in self.bank_items if item.matched_receipt_id == receipt.item_id]

    def selected_preview_tile(self) -> Optional[Dict[str, Any]]:
        for tile in self.preview_tiles:
            if (
                tile.get("kind") == self.selected_attachment_kind
                and tile.get("index") == self.selected_attachment_index
            ):
                return tile
        return None

    def preview_tile_at(self, x: float, y: float) -> Optional[Dict[str, Any]]:
        for tile in reversed(self.preview_tiles):
            left, top, right, bottom = tile.get("bbox", (0, 0, 0, 0))
            if left <= x <= right and top <= y <= bottom:
                return tile
        return None

    def selected_image_state(self) -> Optional[Tuple[Path, Any, int, Optional[Dict[str, Any]], Optional[BankStatementItem]]]:
        tile = self.selected_preview_tile()
        if tile is None:
            return None
        if tile.get("kind") == "proof":
            proof_item = tile.get("proof_item")
            if not isinstance(proof_item, BankStatementItem):
                return None
            return (
                Path(proof_item.path),
                proof_item.crop_box,
                normalize_rotation(proof_item.rotation_degrees),
                None,
                proof_item,
            )
        attachment = None
        receipt = self.selected_item()
        if receipt is not None:
            images = ensure_receipt_images(receipt)
            index = int(tile.get("index") or 0)
            if 0 <= index < len(images):
                attachment = images[index]
        if attachment is None:
            attachment = tile.get("attachment")
        if not isinstance(attachment, dict):
            return None
        return (
            Path(str(attachment.get("path") or "")),
            attachment.get("crop_box"),
            normalize_rotation(attachment.get("rotation_degrees", 0)),
            attachment,
            None,
        )

    def set_selected_image_state(self, crop_box: Any = None, rotation_degrees: Optional[int] = None) -> None:
        state = self.selected_image_state()
        if state is None:
            return
        _path, _crop, _rotation, attachment, proof_item = state
        if proof_item is not None:
            proof_item.crop_box = copy.deepcopy(crop_box)
            if rotation_degrees is not None:
                proof_item.rotation_degrees = normalize_rotation(rotation_degrees)
            return
        if attachment is None:
            return
        receipt = self.selected_item()
        if receipt is not None:
            images = ensure_receipt_images(receipt)
            index = self.selected_attachment_index
            if 0 <= index < len(images):
                attachment = images[index]
        attachment["crop_box"] = copy.deepcopy(crop_box)
        if rotation_degrees is not None:
            attachment["rotation_degrees"] = normalize_rotation(rotation_degrees)
        if receipt is not None:
            ensure_receipt_images(receipt)

    def update_preview(self) -> None:
        if self.preview_canvas is None:
            return
        receipt = self.selected_item()
        if receipt is None:
            self.clear_preview("Select receipt image or PDF files to begin.")
            return
        if Image is None or ImageTk is None:
            self.clear_preview("Image preview is unavailable. Install Pillow from requirements.txt.")
            return
        receipt_images = ensure_receipt_images(receipt)
        show_payment_proof = self.form_version.get() == "USA" and bool(self.bank_items)
        proof_items = self.payment_proofs_for_receipt(receipt) if show_payment_proof else []
        if self.selected_attachment_kind == "proof" and (not proof_items or not show_payment_proof):
            self.selected_attachment_kind = "receipt"
            self.selected_attachment_index = 0
        if self.selected_attachment_kind == "receipt" and self.selected_attachment_index >= len(receipt_images):
            self.selected_attachment_index = 0
        if self.selected_attachment_kind == "proof" and self.selected_attachment_index >= len(proof_items):
            self.selected_attachment_index = 0
        canvas_w = max(240, self.preview_canvas.winfo_width())
        canvas_h = max(240, self.preview_canvas.winfo_height())
        self.preview_canvas.delete("all")
        self.preview_photos = []
        self.preview_tiles = []
        self.preview_action_regions = {}
        self.proof_drop_region = None
        self.photo = None
        self._preview_image_bounds = (0, 0, 0, 0)
        self._preview_original_size = (0, 0)
        self._crop_handle_centers = {}
        try:
            if show_payment_proof:
                gap = 46
                left_w = max(120, int((canvas_w - gap) * 0.56))
                right_w = max(100, canvas_w - left_w - gap)
                proof_x = left_w + gap
                self.proof_drop_region = (proof_x, 0, canvas_w, canvas_h)
                self.draw_preview_section(
                    "Receipt screenshots",
                    0,
                    0,
                    left_w,
                    canvas_h,
                    receipt_images,
                    "receipt",
                )
                self.draw_preview_section(
                    "Payment proof",
                    proof_x,
                    0,
                    right_w,
                    canvas_h,
                    proof_items,
                    "proof",
                )
                self.draw_preview_actions(left_w, gap, canvas_h)
            else:
                self.draw_preview_section(
                    "Receipt screenshots",
                    0,
                    0,
                    canvas_w,
                    canvas_h,
                    receipt_images,
                    "receipt",
                )
            if self.preview_tiles and self.selected_preview_tile() is None:
                first_tile = self.preview_tiles[0]
                self.selected_attachment_kind = str(first_tile.get("kind") or "receipt")
                self.selected_attachment_index = int(first_tile.get("index") or 0)
            self.draw_crop_overlay()
            self.update_preview_buttons()
        except Exception as exc:
            log_exception("Receipt preview failed")
            self.clear_preview(f"Could not preview image:\n{exc}")

    def draw_preview_section(
        self,
        title: str,
        x: int,
        y: int,
        width: int,
        height: int,
        entries: Sequence[Any],
        kind: str,
    ) -> None:
        if self.preview_canvas is None:
            return
        self.preview_canvas.create_text(
            x + 8,
            y + 8,
            text=title,
            anchor="nw",
            fill="#222222",
            font=("Segoe UI", 10, "bold"),
        )
        if kind == "proof":
            self.preview_canvas.create_line(x - 7, y, x - 7, y + height, fill="#D8D8D8")
        if not entries:
            message = "Select Payment Proof, then Generate All." if kind == "proof" else "No receipt screenshot."
            self.preview_canvas.create_text(
                x + width / 2,
                y + height / 2,
                text=message,
                fill="#777777",
                width=max(120, width - 20),
                justify="center",
            )
            return
        columns = 1 if width < 360 or len(entries) == 1 else 2
        rows = max(1, (len(entries) + columns - 1) // columns)
        gap = 10
        top = y + 34
        cell_w = max(72, int((width - gap * (columns - 1) - 8) / columns))
        cell_h = max(86, int((height - 42 - gap * (rows - 1)) / rows))
        for index, entry in enumerate(entries):
            col = index % columns
            row = index // columns
            cell_x = x + 4 + col * (cell_w + gap)
            cell_y = top + row * (cell_h + gap)
            bbox = (cell_x, cell_y, cell_x + cell_w, cell_y + cell_h)
            attachment: Optional[Dict[str, Any]] = None
            proof_item: Optional[BankStatementItem] = None
            if kind == "proof":
                proof_item = entry if isinstance(entry, BankStatementItem) else None
                if proof_item is None:
                    continue
                attachment = attachment_from_item(proof_item)
            else:
                attachment = entry if isinstance(entry, dict) else None
            if attachment is None:
                continue
            path = Path(str(attachment.get("path") or ""))
            is_selected = kind == self.selected_attachment_kind and index == self.selected_attachment_index
            outline = "#0078D7" if is_selected else "#D0D0D0"
            self.preview_canvas.create_rectangle(*bbox, outline=outline, width=2 if is_selected else 1)
            image_bounds = (cell_x + 6, cell_y + 26, 1, 1)
            original_size = (0, 0)
            filename = str(attachment.get("filename") or path.name)
            self.preview_canvas.create_text(
                cell_x + 8,
                cell_y + 6,
                text=filename,
                anchor="nw",
                fill="#444444",
                font=("Segoe UI", 8),
                width=max(60, cell_w - 16),
            )
            if path.exists():
                try:
                    image = oriented_image_from_path(path, attachment.get("rotation_degrees", 0))
                    original_size = (image.width, image.height)
                    max_w = max(1, cell_w - 14)
                    max_h = max(1, cell_h - 34)
                    image.thumbnail((max_w, max_h))
                    photo = ImageTk.PhotoImage(image)
                    self.preview_photos.append(photo)
                    image_x = int(cell_x + (cell_w - image.width) / 2)
                    image_y = int(cell_y + 28 + max(0, (max_h - image.height) / 2))
                    self.preview_canvas.create_image(image_x, image_y, anchor="nw", image=photo)
                    image_bounds = (image_x, image_y, image.width, image.height)
                except Exception:
                    log_exception(f"Could not draw preview tile: {path}")
                    self.preview_canvas.create_text(
                        cell_x + cell_w / 2,
                        cell_y + cell_h / 2,
                        text="Preview failed",
                        fill="#777777",
                    )
            else:
                self.preview_canvas.create_text(
                    cell_x + cell_w / 2,
                    cell_y + cell_h / 2,
                    text="Missing file",
                    fill="#777777",
                )
            self.preview_tiles.append(
                {
                    "kind": kind,
                    "index": index,
                    "bbox": bbox,
                    "image_bounds": image_bounds,
                    "original_size": original_size,
                    "attachment": attachment,
                    "proof_item": proof_item,
                }
            )

    def draw_preview_actions(self, left_width: int, gap: int, canvas_height: int) -> None:
        if self.preview_canvas is None:
            return
        receipt = self.selected_item()
        proofs = self.payment_proofs_for_receipt(receipt) if receipt is not None else []
        has_proofs = bool(proofs)
        can_swap = self.form_version.get() == "USA" and has_proofs and len(self.bank_items) > len(proofs)
        center_x = left_width + gap // 2
        actions = [
            ("swap", "\u21C4", "Swap payment proof", can_swap),
        ]
        start_y = max(56, canvas_height // 2)
        self.preview_action_regions = {}
        for index, (name, icon, label, enabled) in enumerate(actions):
            y = start_y + index * 48
            bbox = (center_x - 17, y - 17, center_x + 17, y + 17)
            fill = "#2563EB" if enabled else "#E2E8F0"
            outline = "#1D4ED8" if enabled else "#CBD5E1"
            text_fill = "#FFFFFF" if enabled else "#94A3B8"
            self.preview_canvas.create_oval(*bbox, fill=fill, outline=outline, width=1)
            self.preview_canvas.create_text(
                center_x,
                y - 1,
                text=icon,
                fill=text_fill,
                font=("Segoe UI Symbol", 14, "bold"),
            )
            self.preview_action_regions[name] = {
                "bbox": bbox,
                "enabled": enabled,
                "label": label,
            }
        if self.proof_drop_region is not None and not has_proofs:
            left, top, right, bottom = self.proof_drop_region
            self.preview_canvas.create_rectangle(
                left + 4,
                top + 34,
                right - 4,
                bottom - 4,
                outline="#CBD5E1",
                dash=(4, 6),
                width=1,
            )

    def preview_action_at(self, x: float, y: float) -> Optional[str]:
        for action, meta in self.preview_action_regions.items():
            left, top, right, bottom = meta.get("bbox", (0, 0, 0, 0))
            if left <= x <= right and top <= y <= bottom:
                return action
        return None

    def point_in_proof_drop_region(self, x: float, y: float) -> bool:
        if self.proof_drop_region is None:
            return False
        left, top, right, bottom = self.proof_drop_region
        return left <= x <= right and top <= y <= bottom

    def run_preview_action(self, action: str) -> None:
        meta = self.preview_action_regions.get(action)
        if not meta or not meta.get("enabled"):
            return
        if action == "swap":
            self.swap_payment_proof()
        elif action == "unlink":
            self.unlink_payment_proof()

    def update_preview_buttons(self) -> None:
        has_tile = self.selected_preview_tile() is not None
        state = "normal" if has_tile else "disabled"
        if not has_tile and self.revert_crop_btn is not None:
            self.revert_crop_btn.configure(state="disabled")
        if self.rotate_left_btn is not None:
            self.rotate_left_btn.configure(state=state)
        if self.rotate_right_btn is not None:
            self.rotate_right_btn.configure(state=state)
        if self.delete_screenshot_btn is not None:
            self.delete_screenshot_btn.configure(state=state)
        receipt = self.selected_item()
        has_proofs = bool(receipt and self.payment_proofs_for_receipt(receipt))
        can_use_proofs = self.form_version.get() == "USA" and receipt is not None and bool(self.bank_items)
        if self.swap_proof_btn is not None:
            self.swap_proof_btn.configure(state="normal" if can_use_proofs else "disabled")
        if self.unlink_proof_btn is not None:
            self.unlink_proof_btn.configure(state="normal" if has_proofs else "disabled")

    def current_preview_crop_points(self) -> List[Tuple[float, float]]:
        state = self.selected_image_state()
        width, height = self._preview_original_size
        crop_box = state[1] if state is not None else None
        crop = normalized_crop_points(crop_box, width, height)
        if crop:
            return crop
        return default_crop_points(width, height)

    def original_to_canvas(self, x: float, y: float) -> Tuple[float, float]:
        image_x, image_y, display_w, display_h = self._preview_image_bounds
        original_w, original_h = self._preview_original_size
        if original_w <= 0 or original_h <= 0:
            return float(image_x), float(image_y)
        return image_x + (x / original_w) * display_w, image_y + (y / original_h) * display_h

    def canvas_to_original(self, x: float, y: float) -> Tuple[float, float]:
        image_x, image_y, display_w, display_h = self._preview_image_bounds
        original_w, original_h = self._preview_original_size
        if display_w <= 0 or display_h <= 0:
            return 0.0, 0.0
        clamped_x = max(image_x, min(image_x + display_w, x))
        clamped_y = max(image_y, min(image_y + display_h, y))
        original_x = ((clamped_x - image_x) / display_w) * original_w
        original_y = ((clamped_y - image_y) / display_h) * original_h
        return original_x, original_y

    def draw_crop_overlay(self) -> None:
        if self.preview_canvas is None:
            return
        tile = self.selected_preview_tile()
        state = self.selected_image_state()
        if tile is None or state is None:
            self.update_preview_buttons()
            return
        self._preview_image_bounds = tuple(tile.get("image_bounds", (0, 0, 0, 0)))  # type: ignore[assignment]
        self._preview_original_size = tuple(tile.get("original_size", (0, 0)))  # type: ignore[assignment]
        original_w, original_h = self._preview_original_size
        image_x, image_y, display_w, display_h = self._preview_image_bounds
        if original_w <= 0 or original_h <= 0 or display_w <= 0 or display_h <= 0:
            self.update_preview_buttons()
            return
        self.preview_canvas.delete("crop")
        crop = self.current_preview_crop_points()
        canvas_points = [self.original_to_canvas(x, y) for x, y in crop]
        flattened_points = [coord for point in canvas_points for coord in point]
        self.preview_canvas.create_polygon(
            *flattened_points,
            outline="#0078D7",
            fill="",
            width=2,
            tags=("crop",),
        )
        handle_size = 8
        centers = {
            "nw": canvas_points[0],
            "ne": canvas_points[1],
            "se": canvas_points[2],
            "sw": canvas_points[3],
        }
        self._crop_handle_centers = centers
        for handle, (hx, hy) in centers.items():
            self.preview_canvas.create_oval(
                hx - handle_size,
                hy - handle_size,
                hx + handle_size,
                hy + handle_size,
                fill="#FFFFFF",
                outline="#0078D7",
                width=2,
                tags=("crop", f"crop_{handle}"),
            )
        if self.revert_crop_btn is not None:
            has_changes = normalized_crop_points(state[1], original_w, original_h) is not None or state[2] != 0
            self.revert_crop_btn.configure(state="normal" if has_changes else "disabled")
        self.preview_canvas.tag_raise("crop")

    def nearest_crop_handle(self, x: float, y: float) -> Optional[str]:
        nearest: Optional[str] = None
        nearest_distance = 16.0
        for handle, (hx, hy) in self._crop_handle_centers.items():
            distance = ((hx - x) ** 2 + (hy - y) ** 2) ** 0.5
            if distance <= nearest_distance:
                nearest = handle
                nearest_distance = distance
        return nearest

    def _on_crop_press(self, event: Any) -> None:
        self._pressed_preview_tile = None
        self._dragging_preview_tile = None
        self._drag_start = (event.x, event.y)
        if self._drag_outline_id is not None and self.preview_canvas is not None:
            self.preview_canvas.delete(self._drag_outline_id)
            self._drag_outline_id = None
        action = self.preview_action_at(event.x, event.y)
        if action is not None:
            self.run_preview_action(action)
            return
        tile = self.preview_tile_at(event.x, event.y)
        if tile is not None:
            kind = str(tile.get("kind") or "receipt")
            index = int(tile.get("index") or 0)
            if kind != self.selected_attachment_kind or index != self.selected_attachment_index:
                self.selected_attachment_kind = kind
                self.selected_attachment_index = index
                self.update_preview()
                tile = self.selected_preview_tile()
            self._pressed_preview_tile = tile
        self._dragging_crop_handle = self.nearest_crop_handle(event.x, event.y)
        if self._dragging_crop_handle is not None:
            self._pressed_preview_tile = None

    def _on_crop_drag(self, event: Any) -> None:
        handle = self._dragging_crop_handle
        original_w, original_h = self._preview_original_size
        if handle is None and self._pressed_preview_tile is not None:
            if self.form_version.get() != "USA" or self._pressed_preview_tile.get("kind") != "receipt":
                return
            dx = event.x - self._drag_start[0]
            dy = event.y - self._drag_start[1]
            if self._dragging_preview_tile is None and (dx * dx + dy * dy) < 64:
                return
            self._dragging_preview_tile = self._pressed_preview_tile
            if self.preview_canvas is not None:
                if self._drag_outline_id is not None:
                    self.preview_canvas.delete(self._drag_outline_id)
                left, top, right, bottom = self._dragging_preview_tile.get("bbox", (0, 0, 0, 0))
                width = right - left
                height = bottom - top
                x1 = int(event.x - width / 2)
                y1 = int(event.y - height / 2)
                self._drag_outline_id = self.preview_canvas.create_rectangle(
                    x1,
                    y1,
                    x1 + width,
                    y1 + height,
                    outline="#2563EB" if self.point_in_proof_drop_region(event.x, event.y) else "#64748B",
                    dash=(5, 4),
                    width=2,
                )
                self.preview_canvas.tag_raise(self._drag_outline_id)
            self.status_text.set("Release in Payment proof to move the screenshot.")
            return
        if self.selected_image_state() is None or handle is None or original_w <= 0 or original_h <= 0:
            return
        points = self.current_preview_crop_points()
        x, y = self.canvas_to_original(event.x, event.y)
        handle_indexes = {"nw": 0, "ne": 1, "se": 2, "sw": 3}
        points[handle_indexes[handle]] = (x, y)
        state = self.selected_image_state()
        rotation = state[2] if state is not None else 0
        self.set_selected_image_state(flatten_crop_points(points), rotation)
        self.draw_crop_overlay()

    def _on_crop_release(self, event: Any) -> None:
        if self._dragging_preview_tile is not None:
            tile = self._dragging_preview_tile
            dropped = self.point_in_proof_drop_region(event.x, event.y)
            if self._drag_outline_id is not None and self.preview_canvas is not None:
                self.preview_canvas.delete(self._drag_outline_id)
            self._drag_outline_id = None
            self._dragging_preview_tile = None
            self._pressed_preview_tile = None
            if dropped:
                self.move_receipt_screenshot_to_payment_proof(tile)
            else:
                self.status_text.set("Screenshot move cancelled.")
            return
        self._pressed_preview_tile = None
        if self._dragging_crop_handle is None:
            return
        self._dragging_crop_handle = None
        self.save_session()
        self.status_text.set("Crop updated for selected screenshot.")

    def revert_crop(self) -> None:
        state = self.selected_image_state()
        if state is None:
            return
        self.set_selected_image_state(None, 0)
        self.update_preview()
        self.save_session()
        self.status_text.set("Selected screenshot reverted to original.")

    def rotate_selected(self, delta_degrees: int) -> None:
        state = self.selected_image_state()
        if state is None:
            return
        path, crop_box, rotation_degrees, _attachment, _proof_item = state
        try:
            image = oriented_image_from_path(path, rotation_degrees)
            transformed_crop = rotated_crop_points(
                crop_box,
                image.width,
                image.height,
                delta_degrees,
            )
        except Exception:
            log_exception("Could not transform crop during rotation")
            transformed_crop = None
        new_rotation = normalize_rotation(normalize_rotation(rotation_degrees) + delta_degrees)
        self.set_selected_image_state(transformed_crop, new_rotation)
        self.update_preview()
        self.save_session()
        direction = "right" if delta_degrees > 0 else "left"
        self.status_text.set(f"Rotated selected screenshot {direction}.")

    def move_receipt_screenshot_to_payment_proof(self, tile: Dict[str, Any]) -> None:
        if self.form_version.get() != "USA":
            return
        receipt = self.selected_item()
        if receipt is None:
            return
        index = int(tile.get("index") or 0)
        images = ensure_receipt_images(receipt)
        if index < 0 or index >= len(images):
            return
        attachment = copy.deepcopy(images[index])
        filename = str(attachment.get("filename") or Path(str(attachment.get("path") or "")).name)
        removing_row = len(images) == 1
        if removing_row and messagebox and not messagebox.askyesno(
            "Move to payment proof",
            f"Move this screenshot to Payment proof and remove the receipt row?\n\n{filename}",
        ):
            return
        proof_item = BankStatementItem(
            item_id=uuid.uuid4().hex,
            path=str(attachment.get("path") or ""),
            filename=filename,
            source_path=str(attachment.get("source_path") or ""),
            source_page=str(attachment.get("source_page") or ""),
            date=receipt.date,
            amount=receipt.amount,
            place=receipt.place,
            matched_receipt_id="" if removing_row else receipt.item_id,
            status="Needs manual review" if removing_row else "Matched manually",
            crop_box=copy.deepcopy(attachment.get("crop_box")),
            rotation_degrees=normalize_rotation(attachment.get("rotation_degrees", 0)),
        )
        self.bank_items.append(proof_item)
        del images[index]
        if not removing_row:
            receipt.receipt_images = images
            ensure_receipt_images(receipt)
            self.selected_attachment_kind = "proof"
            self.selected_attachment_index = len(self.payment_proofs_for_receipt(receipt)) - 1
            self.refresh_tree()
            self.refresh_bank_tree()
            self.update_preview()
            self.save_session()
            self.status_text.set(f"Moved {filename} to Payment proof.")
            self._details_ready_for_export = False
            self.update_toolbar_recommendation()
            return

        remove_index = self.selected_index
        if remove_index is not None and 0 <= remove_index < len(self.items):
            del self.items[remove_index]
        self.selected_index = None
        self.selected_attachment_kind = "receipt"
        self.selected_attachment_index = 0
        self.refresh_tree()
        self.refresh_bank_tree()
        if self.items:
            self.select_index(min(remove_index or 0, len(self.items) - 1))
        else:
            self.load_selected_into_fields()
            self.update_preview()
        self.save_session()
        self.status_text.set("Moved screenshot to Payment proof. It needs manual review.")
        self._details_ready_for_export = False
        self.update_toolbar_recommendation()

    def delete_selected_screenshot(self) -> None:
        tile = self.selected_preview_tile()
        receipt = self.selected_item()
        if tile is None or receipt is None:
            return
        kind = str(tile.get("kind") or "receipt")
        index = int(tile.get("index") or 0)
        if kind == "proof":
            proof_item = tile.get("proof_item")
            if not isinstance(proof_item, BankStatementItem):
                return
            if messagebox and not messagebox.askyesno(
                "Delete screenshot",
                f"Remove this payment proof image?\n\n{proof_item.filename}",
            ):
                return
            self.bank_items = [item for item in self.bank_items if item.item_id != proof_item.item_id]
            self.selected_attachment_kind = "receipt"
            self.selected_attachment_index = 0
            self.refresh_bank_tree()
            self.update_preview()
            self.save_session()
            self.status_text.set("Deleted selected payment proof screenshot.")
            self._details_ready_for_export = False
            self.update_toolbar_recommendation()
            return

        images = ensure_receipt_images(receipt)
        if index < 0 or index >= len(images):
            return
        filename = str(images[index].get("filename") or Path(str(images[index].get("path") or "")).name)
        if messagebox and not messagebox.askyesno(
            "Delete screenshot",
            f"Remove this receipt screenshot?\n\n{filename}",
        ):
            return
        del images[index]
        if images:
            receipt.receipt_images = images
            ensure_receipt_images(receipt)
            self.selected_attachment_index = max(0, min(index, len(images) - 1))
            self.refresh_tree()
            self.update_preview()
            self.save_session()
            self.status_text.set("Deleted selected receipt screenshot.")
            return

        removed_id = receipt.item_id
        remove_index = self.selected_index
        if remove_index is not None and 0 <= remove_index < len(self.items):
            del self.items[remove_index]
        for proof_item in self.bank_items:
            if proof_item.matched_receipt_id == removed_id:
                proof_item.matched_receipt_id = ""
                proof_item.status = "Needs manual review"
        self.selected_index = None
        self.refresh_tree()
        if self.items:
            self.select_index(min(remove_index or 0, len(self.items) - 1))
        else:
            self.load_selected_into_fields()
            self.update_preview()
        self.save_session()
        self.status_text.set("Deleted receipt row because its last screenshot was removed.")

    def swap_payment_proof(self) -> None:
        receipt = self.selected_item()
        if receipt is None:
            return
        if self.form_version.get() != "USA":
            self.show_info("Payment proof is only used for the USA form.")
            return
        if not self.bank_items:
            self.show_info("Select payment proof files first.")
            return
        current_indexes = [
            index for index, proof_item in enumerate(self.bank_items)
            if proof_item.matched_receipt_id == receipt.item_id
        ]
        start = current_indexes[0] if current_indexes else -1
        chosen_index: Optional[int] = None
        for offset in range(1, len(self.bank_items) + 1):
            candidate_index = (start + offset) % len(self.bank_items)
            if candidate_index not in current_indexes:
                chosen_index = candidate_index
                break
        if chosen_index is None:
            self.show_info("No other payment proof is loaded.")
            return
        for index in current_indexes:
            self.bank_items[index].matched_receipt_id = ""
            self.bank_items[index].status = "Needs manual review"
        chosen = self.bank_items[chosen_index]
        chosen.matched_receipt_id = receipt.item_id
        chosen.status = "Matched manually"
        self.selected_attachment_kind = "proof"
        self.selected_attachment_index = 0
        self.refresh_bank_tree()
        self.update_preview()
        self.save_session()
        self.status_text.set(f"Swapped payment proof to {chosen.filename}.")
        self._details_ready_for_export = False
        self.update_toolbar_recommendation()

    def unlink_payment_proof(self) -> None:
        receipt = self.selected_item()
        if receipt is None:
            return
        changed = 0
        for proof_item in self.bank_items:
            if proof_item.matched_receipt_id == receipt.item_id:
                proof_item.matched_receipt_id = ""
                proof_item.status = "Needs manual review"
                changed += 1
        if changed:
            self.selected_attachment_kind = "receipt"
            self.selected_attachment_index = 0
            self.refresh_bank_tree()
            self.update_preview()
            self.save_session()
            self.status_text.set("Unlinked payment proof from selected receipt.")
            self._details_ready_for_export = False
            self.update_toolbar_recommendation()

    def remove_selected(self, _event: Any = None) -> str:
        indices = self.selected_indices()
        if not indices:
            return "break"
        if messagebox and not messagebox.askyesno(
            "Remove receipts",
            f"Remove {len(indices)} selected receipt row(s)?",
        ):
            return "break"
        removed_ids = {self.items[index].item_id for index in indices if 0 <= index < len(self.items)}
        for index in sorted(indices, reverse=True):
            del self.items[index]
        for bank_item in self.bank_items:
            if bank_item.matched_receipt_id in removed_ids:
                bank_item.matched_receipt_id = ""
                bank_item.status = "Needs manual review"
        self.selected_index = None
        self.refresh_tree()
        self.refresh_bank_tree()
        if self.items:
            next_index = min(indices[0], len(self.items) - 1)
            self._selection_anchor_index = next_index
            self.select_index(next_index)
        else:
            self._selection_anchor_index = None
            self.load_selected_into_fields()
            self.update_preview()
        self.status_text.set(f"Removed {len(indices)} selected receipt(s).")
        self._details_ready_for_export = False
        self.update_toolbar_recommendation()
        self.save_session()
        return "break"

    def clear_all(self) -> None:
        if not self.items and not self.bank_items and not self.exchange_rate_items:
            return
        if messagebox and not messagebox.askyesno("Clear receipts", "Remove all uploaded receipt rows and support images?"):
            return
        self.items.clear()
        self.bank_items.clear()
        self.exchange_rate_items.clear()
        self.selected_index = None
        self.selected_bank_index = None
        self._selection_anchor_index = None
        self.refresh_tree()
        self.refresh_bank_tree()
        self.load_selected_into_fields()
        self.update_preview()
        self.status_text.set("Cleared receipt list.")
        self._details_ready_for_export = False
        self.update_toolbar_recommendation()
        self.save_session()

    def apply_ai_data_to_item(self, item: ReceiptItem, data: Dict[str, Any]) -> None:
        field_map = {
            "date": "date",
            "place": "place",
            "vendor": "place",
            "amount": "amount",
            "currency": "currency",
            "krw_amount": "krw_amount",
            "amount_krw": "krw_amount",
            "rmb_amount": "rmb_amount",
            "amount_rmb": "rmb_amount",
            "purpose": "purpose",
            "details": "details",
            "project_number": "project_number",
            "category": "category",
            "payment_method": "payment_method",
            "receipt_label": "receipt_label",
        }
        for src, dest in field_map.items():
            value = data.get(src)
            if value is None:
                continue
            text = str(value).strip()
            if not text:
                continue
            if dest == "category":
                text = normalized_category(text, self.form_version.get())
            elif dest == "currency":
                text = normalize_currency(text, "USD" if self.form_version.get() == "USA" else item.currency or "KRW")
            elif dest == "amount":
                hinted_currency = normalize_currency(text, "")
                if hinted_currency:
                    item.currency = hinted_currency
            setattr(item, dest, text)
        if self.form_version.get() == "USA":
            item.currency = "USD"
        else:
            item.currency = normalize_currency(item.currency, "KRW")
        self.update_item_amount_fields(item, "amount")
        item.status = "AI filled"

    def apply_ai_data_to_bank_item(self, item: BankStatementItem, data: Dict[str, Any]) -> None:
        for source_key, dest_key in {
            "date": "date",
            "amount": "amount",
            "place": "place",
            "vendor": "place",
            "details": "place",
        }.items():
            value = data.get(source_key)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                setattr(item, dest_key, text)
        item.status = "Needs match"

    def same_expense_date(self, left: str, right: str) -> bool:
        left_value = parse_date_value(left)
        right_value = parse_date_value(right)
        if isinstance(left_value, date) and isinstance(right_value, date):
            return left_value == right_value
        return (left or "").strip() == (right or "").strip()

    def match_bank_item(self, bank_item: BankStatementItem) -> None:
        bank_amount = safe_float(bank_item.amount)
        if bank_amount is None or not bank_item.date:
            bank_item.matched_receipt_id = ""
            bank_item.status = "Needs manual review"
            return
        matches: List[ReceiptItem] = []
        for receipt in self.items:
            receipt_amount = safe_float(receipt.amount)
            if receipt_amount is None:
                continue
            if abs(receipt_amount - bank_amount) > 0.01:
                continue
            if not self.same_expense_date(receipt.date, bank_item.date):
                continue
            matches.append(receipt)
        if len(matches) == 1:
            bank_item.matched_receipt_id = matches[0].item_id
            bank_item.status = "Matched"
        else:
            bank_item.matched_receipt_id = ""
            bank_item.status = "Needs manual review"

    def match_all_bank_items(self) -> None:
        for bank_item in self.bank_items:
            self.match_bank_item(bank_item)
        self.resolve_payment_proof_conflicts()
        self.refresh_bank_tree()

    def resolve_payment_proof_conflicts(self) -> None:
        grouped: Dict[str, List[BankStatementItem]] = {}
        for bank_item in self.bank_items:
            if bank_item.matched_receipt_id:
                grouped.setdefault(bank_item.matched_receipt_id, []).append(bank_item)
        for proofs in grouped.values():
            if len(proofs) <= 1:
                continue
            keeper = proofs[0]
            keeper.status = "Matched"
            for proof_item in proofs[1:]:
                proof_item.matched_receipt_id = ""
                proof_item.status = "Needs manual review"

    def receipt_merge_key(self, item: ReceiptItem) -> Optional[Tuple[str, float]]:
        amount = safe_float(item.amount)
        if amount is None:
            return None
        date_value = parse_date_value(item.date)
        if isinstance(date_value, date):
            date_key = date_value.isoformat()
        else:
            date_key = (item.date or "").strip()
        if not date_key:
            return None
        return date_key, round(amount, 2)

    def merge_same_usa_receipts(self) -> int:
        if self.form_version.get() != "USA":
            return 0
        merged_count = 0
        merged_items: List[ReceiptItem] = []
        seen: Dict[Tuple[str, float], ReceiptItem] = {}
        removed_to_target: Dict[str, str] = {}
        for item in self.items:
            ensure_receipt_images(item)
            key = self.receipt_merge_key(item)
            if key is None or key not in seen:
                if key is not None:
                    seen[key] = item
                merged_items.append(item)
                continue
            target = seen[key]
            merge_attachment_lists(ensure_receipt_images(target), ensure_receipt_images(item))
            for attr in (
                "place",
                "purpose",
                "details",
                "project_number",
                "category",
                "payment_method",
                "receipt_label",
                "rmb_amount",
            ):
                if not str(getattr(target, attr, "") or "").strip() and str(getattr(item, attr, "") or "").strip():
                    setattr(target, attr, getattr(item, attr))
            if target.status not in {"Failed", "Processing"}:
                target.status = "Merged"
            removed_to_target[item.item_id] = target.item_id
            merged_count += 1
        if not merged_count:
            return 0
        self.items = merged_items
        for proof_item in self.bank_items:
            if proof_item.matched_receipt_id in removed_to_target:
                proof_item.matched_receipt_id = removed_to_target[proof_item.matched_receipt_id]
        if self.selected_index is not None and self.selected_index >= len(self.items):
            self.selected_index = max(0, len(self.items) - 1) if self.items else None
        return merged_count

    def generate_selected_details(self) -> None:
        self.save_current_fields()
        items = self.selected_items()
        if not items:
            self.show_info("Upload and select a receipt first.")
            return
        self._run_ai_for_items(items, reload_selected=True)

    def generate_all_details(self) -> None:
        self.save_current_fields()
        if not self.items:
            self.show_info("Upload receipts first.")
            return
        self._run_ai_for_items(list(self.items), reload_selected=True, include_bank=True)

    def _run_ai_for_items(
        self,
        items: List[ReceiptItem],
        reload_selected: bool,
        include_bank: bool = False,
    ) -> None:
        if self._busy:
            self.show_info("A batch is already running.")
            return
        bank_items = list(self.bank_items) if include_bank and self.form_version.get() == "USA" else []
        total = len(items) + len(bank_items)
        if total <= 0:
            return
        form_version = self.form_version.get()
        if include_bank:
            self._details_ready_for_export = False
        self.set_busy(True)
        self.set_progress(0, total, f"0/{total}")
        self.status_text.set(f"Generating details for {total} image(s)...")

        def worker() -> None:
            successes = 0
            failures: List[str] = []
            for idx, item in enumerate(items, start=1):
                self.root.after(
                    0,
                    lambda i=item, n=idx: self.mark_item_status(i, f"Processing {n}/{total}", n - 1, total),
                )
                self.set_status(f"Generating details {idx}/{total}: {item.filename}")
                try:
                    data = call_openai_receipt_extraction(
                        Path(item.path),
                        form_version,
                        item,
                        progress=self.set_status,
                    )
                    self.apply_ai_data_to_item(item, data)
                    self.move_item_to_processed(item)
                    successes += 1
                    self.root.after(
                        0,
                        lambda i=item, n=idx: self.mark_item_status(i, "AI filled", n, total),
                    )
                except Exception as exc:
                    setup_logging()
                    LOGGER.exception("AI extraction failed for %s", item.filename)
                    failures.append(f"{item.filename}: {exc}")
                    self.root.after(
                        0,
                        lambda i=item, n=idx: self.mark_item_status(i, "Failed", n, total),
                    )
                    continue
            if include_bank and form_version == "USA":
                merged_count = self.merge_same_usa_receipts()
                if merged_count:
                    self.root.after(0, self.refresh_tree)
                    self.set_status(f"Merged {merged_count} duplicate receipt screenshot(s).")
            offset = len(items)
            for bank_index, bank_item in enumerate(bank_items, start=1):
                progress_index = offset + bank_index
                self.root.after(
                    0,
                    lambda i=bank_item, n=progress_index: self.mark_bank_item_status(
                        i, f"Processing {n}/{total}", n - 1, total
                    ),
                )
                self.set_status(f"Reading payment proof {bank_index}/{len(bank_items)}: {bank_item.filename}")
                try:
                    temp = ReceiptItem(
                        item_id=bank_item.item_id,
                        path=bank_item.path,
                        filename=bank_item.filename,
                        source_path=bank_item.source_path,
                        source_page=bank_item.source_page,
                        date=bank_item.date,
                        place=bank_item.place,
                        amount=bank_item.amount,
                        currency="USD",
                        crop_box=bank_item.crop_box,
                        rotation_degrees=bank_item.rotation_degrees,
                    )
                    data = call_openai_receipt_extraction(
                        Path(bank_item.path),
                        "USA",
                        temp,
                        progress=self.set_status,
                    )
                    self.apply_ai_data_to_bank_item(bank_item, data)
                    self.match_bank_item(bank_item)
                    successes += 1
                    self.root.after(
                        0,
                        lambda i=bank_item, n=progress_index: self.mark_bank_item_status(i, i.status, n, total),
                    )
                except Exception as exc:
                    setup_logging()
                    LOGGER.exception("Bank statement extraction failed for %s", bank_item.filename)
                    bank_item.status = "Needs manual review"
                    failures.append(f"{bank_item.filename}: {exc}")
                    self.root.after(
                        0,
                        lambda i=bank_item, n=progress_index: self.mark_bank_item_status(
                            i, "Needs manual review", n, total
                        ),
                    )
                    continue
            if bank_items:
                self.resolve_payment_proof_conflicts()
            self.root.after(0, lambda: self.after_ai_complete(successes, failures, total, reload_selected, include_bank))

        threading.Thread(target=worker, daemon=True).start()

    def after_ai_complete(
        self,
        successes: int,
        failures: List[str],
        total: int,
        reload_selected: bool,
        ready_for_export: bool = False,
    ) -> None:
        self.set_busy(False)
        if ready_for_export and successes > 0:
            self._details_ready_for_export = True
        self.set_progress(total, total, f"{total}/{total}")
        if ready_for_export and successes > 0:
            self.sort_items_for_current_report()
        else:
            self.refresh_tree()
        self.refresh_bank_tree()
        if reload_selected:
            self.load_selected_into_fields()
            self.update_preview()
        self.save_session()
        if failures:
            self.status_text.set(f"AI finished: {successes} succeeded, {len(failures)} failed.")
            if messagebox:
                shown = "\n".join(failures[:5])
                if len(failures) > 5:
                    shown += f"\n...and {len(failures) - 5} more."
                messagebox.showwarning(
                    "Generate Details",
                    f"Finished with failures.\n\n{shown}\n\nLogged to:\n{LOG_DIR / 'app.log'}",
                )
        else:
            self.status_text.set("AI details generated. Review before exporting.")
        self.update_toolbar_recommendation()

    def set_status(self, message: str) -> None:
        self.root.after(0, lambda: self.status_text.set(message))

    def set_busy(self, busy: bool) -> None:
        self._busy = busy
        state = "disabled" if busy else "normal"
        for button in (
            self.upload_folder_btn,
            self.generate_details_btn,
            self.generate_all_btn,
            self.generate_excel_btn,
        ):
            try:
                button.configure(state=state)
            except Exception:
                pass
        if self.select_payment_proof_btn is not None:
            proof_state = "normal" if not busy else "disabled"
            try:
                self.select_payment_proof_btn.configure(state=proof_state)
            except Exception:
                pass
        self.update_toolbar_recommendation()

    def set_progress(self, done: int, total: int, text: str = "") -> None:
        if total <= 0:
            self.progress_value.set(0)
            self.progress_text.set(text)
            return
        pct = max(0.0, min(100.0, (done / total) * 100.0))
        self.progress_value.set(pct)
        self.progress_text.set(text or f"{done}/{total}")

    def mark_item_status(self, item: ReceiptItem, status: str, done: int, total: int) -> None:
        item.status = status
        self.refresh_tree()
        self.set_progress(done, total, f"{done}/{total}")
        if self.selected_item() is item:
            self.load_selected_into_fields()
            self.update_preview()

    def mark_bank_item_status(self, item: BankStatementItem, status: str, done: int, total: int) -> None:
        item.status = status
        self.refresh_bank_tree()
        self.set_progress(done, total, f"{done}/{total}")

    def move_item_to_processed(self, item: ReceiptItem) -> bool:
        source = Path(item.path)
        if not source.exists() or not same_folder(source, UNPROCESSED_DIR):
            return False
        try:
            ensure_runtime_folders()
            destination = unique_path(PROCESSED_DIR, source.name)
            shutil.move(str(source), str(destination))
            item.path = str(destination)
            item.filename = destination.name
            LOGGER.info("Moved processed receipt to %s", destination)
            return True
        except Exception:
            log_exception(f"Could not move receipt to Processed: {source}")
            return False

    def move_all_loaded_to_processed(self) -> None:
        moved = False
        for item in self.items:
            moved = self.move_item_to_processed(item) or moved
        if moved:
            self.refresh_tree()
            self.update_preview()

    def generate_excel(self) -> None:
        self.save_current_fields()
        if not self.items:
            self.show_info("Upload at least one receipt before generating Excel.")
            return
        if filedialog is None:
            return
        default_name = f"reimbursement_{self.form_version.get().lower()}_{date.today().strftime('%Y%m%d')}.xlsx"
        output = filedialog.asksaveasfilename(
            title="Save reimbursement workbook",
            initialdir=str(OUTPUT_DIR),
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not output:
            return
        output_path = Path(output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            if self.form_version.get() == "USA":
                exchange_rate = safe_float(self.exchange_rate.get()) or safe_float(DEFAULT_USD_TO_RMB_RATE) or 6.8
                export_usa(self.items, output_path, exchange_rate, self.bank_items)
            else:
                krw_rate = safe_float(self.krw_to_rmb_rate.get()) or safe_float(DEFAULT_KRW_TO_RMB_RATE) or 0.0046
                usd_rate = safe_float(self.exchange_rate.get()) or safe_float(DEFAULT_USD_TO_RMB_RATE) or 6.8
                usd_krw_rate = safe_float(self.usd_to_krw_rate.get()) or safe_float(DEFAULT_USD_TO_KRW_RATE) or 1548.86
                export_korea(self.items, output_path, krw_rate, usd_rate, usd_krw_rate, self.exchange_rate_items)
        except Exception as exc:
            log_path = log_exception("Excel export failed")
            if messagebox:
                messagebox.showerror(
                    "Generate Excel",
                    f"Could not generate workbook:\n{exc}\n\nLogged to:\n{log_path}",
                )
            self.status_text.set("Export failed.")
            return
        self.last_generated_output_path = output_path
        self.move_all_loaded_to_processed()
        self.save_session()
        self.status_text.set(f"Saved: {output_path}")
        if messagebox:
            choice = messagebox.askyesnocancel(
                "Generate Excel",
                "Workbook saved.\n\n"
                "Yes: open the Excel file\n"
                "No: show it in the folder\n"
                "Cancel: stay here",
            )
            try:
                if choice is True:
                    open_path(output_path)
                elif choice is False:
                    reveal_in_file_explorer(output_path)
            except Exception:
                log_exception("Could not open generated workbook or folder")

    def open_output_folder(self) -> None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        target = self.last_generated_output_path
        if target is None or not target.exists():
            workbooks = list(OUTPUT_DIR.glob("*.xlsx"))
            target = max(workbooks, key=lambda path: path.stat().st_mtime, default=None)
        if target is not None and target.exists():
            try:
                reveal_in_file_explorer(target)
                return
            except Exception:
                log_exception("Could not reveal output workbook")
        if sys.platform == "win32":
            os.startfile(str(OUTPUT_DIR))  # type: ignore[attr-defined]
        else:
            self.show_info(f"Output folder: {OUTPUT_DIR}")

    def show_info(self, message: str) -> None:
        if messagebox:
            messagebox.showinfo("Reimbursement Helper", message)
        else:
            print(message)

    def session_payload(self) -> Dict[str, Any]:
        self.save_current_fields()
        return {
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            "form_version": self.form_version.get(),
            "usa_exchange_rate": self.exchange_rate.get(),
            "usd_to_krw_rate": self.usd_to_krw_rate.get(),
            "krw_to_rmb_rate": self.krw_to_rmb_rate.get(),
            "selected_index": self.selected_index,
            "items": [asdict(item) for item in self.items],
            "bank_items": [asdict(item) for item in self.bank_items],
            "exchange_rate_items": [asdict(item) for item in self.exchange_rate_items],
        }

    def save_session(self) -> None:
        try:
            CONFIG_DIR.mkdir(parents=True, exist_ok=True)
            SESSION_FILE.write_text(
                json.dumps(self.session_payload(), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception:
            log_exception("Could not save session state")

    def restore_previous_session_if_available(self) -> None:
        if not SESSION_FILE.exists():
            return
        try:
            data = json.loads(SESSION_FILE.read_text(encoding="utf-8"))
        except Exception:
            log_exception("Could not read previous session state")
            return
        raw_items = data.get("items")
        raw_bank_items = data.get("bank_items")
        raw_exchange_rate_items = data.get("exchange_rate_items")
        if not isinstance(raw_items, list):
            raw_items = []
        if not isinstance(raw_bank_items, list):
            raw_bank_items = []
        if not isinstance(raw_exchange_rate_items, list):
            raw_exchange_rate_items = []
        if not raw_items and not raw_bank_items and not raw_exchange_rate_items:
            return
        should_restore = True
        if messagebox:
            saved_at = str(data.get("saved_at") or "last time")
            should_restore = messagebox.askyesno(
                "Restore Previous File",
                f"A previous reimbursement session was saved at {saved_at}.\n\nRestore it?",
            )
        if not should_restore:
            self.status_text.set("Started a fresh reimbursement session.")
            return
        fields = set(ReceiptItem.__dataclass_fields__.keys())
        restored: List[ReceiptItem] = []
        for raw in raw_items:
            if not isinstance(raw, dict):
                continue
            values = {key: raw.get(key, "") for key in fields}
            if not values.get("item_id"):
                values["item_id"] = uuid.uuid4().hex
            if not isinstance(values.get("crop_box"), list):
                values["crop_box"] = None
            if not isinstance(values.get("receipt_images"), list):
                values["receipt_images"] = []
            values["rotation_degrees"] = normalize_rotation(values.get("rotation_degrees", 0))
            item = ReceiptItem(**values)
            ensure_receipt_images(item)
            restored.append(item)
        if not restored:
            restored = []
        bank_fields = set(BankStatementItem.__dataclass_fields__.keys())
        restored_bank: List[BankStatementItem] = []
        for raw in raw_bank_items:
            if not isinstance(raw, dict):
                continue
            values = {key: raw.get(key, "") for key in bank_fields}
            if not values.get("item_id"):
                values["item_id"] = uuid.uuid4().hex
            if not isinstance(values.get("crop_box"), list):
                values["crop_box"] = None
            values["rotation_degrees"] = normalize_rotation(values.get("rotation_degrees", 0))
            restored_bank.append(BankStatementItem(**values))
        restored_exchange: List[BankStatementItem] = []
        for raw in raw_exchange_rate_items:
            if not isinstance(raw, dict):
                continue
            values = {key: raw.get(key, "") for key in bank_fields}
            if not values.get("item_id"):
                values["item_id"] = uuid.uuid4().hex
            if not isinstance(values.get("crop_box"), list):
                values["crop_box"] = None
            values["rotation_degrees"] = normalize_rotation(values.get("rotation_degrees", 0))
            restored_exchange.append(BankStatementItem(**values))
        self.form_version.set(normalized_form_version(data.get("form_version") or self.form_version.get()))
        self.exchange_rate.set(str(data.get("usa_exchange_rate") or DEFAULT_USD_TO_RMB_RATE))
        self.usd_to_krw_rate.set(str(data.get("usd_to_krw_rate") or DEFAULT_USD_TO_KRW_RATE))
        self.krw_to_rmb_rate.set(str(data.get("krw_to_rmb_rate") or DEFAULT_KRW_TO_RMB_RATE))
        self.items = restored
        self.bank_items = restored_bank
        self.exchange_rate_items = restored_exchange
        self._last_form_version = self.form_version.get()
        try:
            selected = int(data.get("selected_index", 0) or 0)
        except (TypeError, ValueError):
            selected = 0
        self.selected_index = None
        self._on_form_version_changed()
        self.refresh_tree()
        self.refresh_bank_tree()
        if self.items:
            self.select_index(max(0, min(selected, len(self.items) - 1)))
        self.status_text.set("Previous reimbursement session restored.")

    def on_close(self) -> None:
        self.settings["usa_exchange_rate"] = self.exchange_rate.get()
        self.settings["usd_to_krw_rate"] = self.usd_to_krw_rate.get()
        self.settings["krw_to_rmb_rate"] = self.krw_to_rmb_rate.get()
        self.settings["last_form_version"] = normalized_form_version(self.form_version.get())
        self.settings["language"] = normalize_language(self.language.get())
        save_user_settings(self.settings)
        self.save_session()
        self.root.destroy()


def smoke_test_export() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    sample = ReceiptItem(
        item_id="sample",
        path=str(APP_DIR / "examples" / "sample_receipt_placeholder.png"),
        filename="sample_receipt_placeholder.png",
        date=date.today().strftime("%Y-%m-%d"),
        place="Sample Vendor",
        amount="12.34",
        currency="USD",
        krw_amount="18000",
        purpose="Parking",
        details="Sample parking receipt",
        project_number="",
        category="transportation",
        payment_method="",
        receipt_label="Sample receipt",
        status="Sample",
    )
    if Image is not None:
        sample_path = Path(sample.path)
        sample_path.parent.mkdir(parents=True, exist_ok=True)
        if not sample_path.exists():
            img = Image.new("RGB", (360, 220), "#FFFFFF")
            img.save(sample_path)
    export_usa([sample], OUTPUT_DIR / "smoke_usa.xlsx", safe_float(DEFAULT_USD_TO_RMB_RATE) or 6.8)
    export_korea(
        [sample],
        OUTPUT_DIR / "smoke_korea.xlsx",
        safe_float(DEFAULT_KRW_TO_RMB_RATE) or 0.0046,
        safe_float(DEFAULT_USD_TO_RMB_RATE) or 6.8,
        safe_float(DEFAULT_USD_TO_KRW_RATE) or 1548.86,
    )
    print("Smoke exports created.")


def main(argv: Optional[List[str]] = None) -> int:
    args = list(sys.argv[1:] if argv is None else argv)
    try:
        setup_logging()
        ensure_dependencies()
        if "--smoke-test" in args:
            smoke_test_export()
            return 0
        ensure_runtime_folders()
        root = tk.Tk()
        ReimbursementHelperApp(root)
        root.mainloop()
        return 0
    except Exception as exc:
        log_path = log_exception("Application startup failed")
        details = "".join(traceback.format_exception_only(type(exc), exc)).strip()
        if messagebox and tk is not None:
            try:
                root = tk.Tk()
                root.withdraw()
                messagebox.showerror("Reimbursement Helper", f"{details}\n\nLogged to:\n{log_path}")
                root.destroy()
            except Exception:
                print(details)
        else:
            print(details)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
